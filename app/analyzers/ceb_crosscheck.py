"""CEB 3자 수량 교차검증 — 성적서 × CEB 청구 × NIS 지급물량표.

수량 단위 사슬 (2026-07-06 파일럿 실증):
    성적서(부재유형별 수량, pcs) × CEB(부재당 용접점 unit Q'ty) = 물량표 VT(점) = CEB 청구 VT(점)

검증 (과다 중심 — 미청구·미지급은 정보성):
  ① 성적서 ↔ 물량표: report_no 조인. 물량표에 없는 성적서 = NIS 미지급 후보 (지급의무 룰 참고 목록)
  ② 수량 대사: Σ(성적서 부재수 × CEB unit Q'ty) vs 물량표 VT — 불일치 목록
  ③ 부재수 대사: Σ(성적서 부재수 by 건물·부재유형) vs CEB DD Q'ty — 초과 검사분 표시
  ④ 총량: Σ물량표 vs Σ CEB 청구 (기존 확인 0.07%)

주의: 보유 성적서가 일부(mac 표본)일 수 있으므로 "성적서 미보유" 는 발견사항 아님 (커버리지 명시).
"""
from __future__ import annotations

import json
import logging
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from app.extractors.civil_report_parser import (
    CivilReport,
    normalize_product,
    normalize_report_no,
)

logger = logging.getLogger(__name__)


def normalize_building(s: Optional[str]) -> str:
    """'9.1UGG'/'91ULD'/'20UMA' → 점 제거 대문자."""
    return re.sub(r"[.\s]", "", str(s or "").upper())


def building_from_drawing(drawing: Optional[str]) -> Optional[str]:
    """도면번호에서 건물 유도 — MD.D.N000.2.0UMA95... → '20UMA'."""
    if not drawing:
        return None
    m = re.search(r"MD\.[DС]\.[A-Z0-9]{4}\.(\d)\.([0-9][A-Z]{3})", drawing)
    if not m:
        return None
    unit, bldg = m.group(1), m.group(2)   # bldg 예: '0UMA'
    return f"{unit}{bldg}"                 # '20UMA'


# ─────────────────────────── 로더 ───────────────────────────


def load_nis_ledger(xlsx_path: Path) -> dict[str, dict]:
    """NIS 지급물량표 → {report_no_norm: {sheet, vt, ut, round, date}}.

    한 셀에 복수 성적서(개행·쉼표) 기재 가능 → 분해. 소계/총계 행(성적서 없음) 제외.
    """
    import openpyxl

    ledger: dict[str, dict] = {}
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    for sheet in ("CP-A1", "CP-E1", "CP-P1"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=3, values_only=True):
            raw = str(row[2] or "")
            toks = [normalize_report_no(t) for t in re.split(r"[\n,;]", raw)]
            toks = [t for t in toks if t and t.lower() != "none"]
            if not toks:
                continue
            for t in toks:
                ledger[t] = {
                    "sheet": sheet,
                    "vt": _num(row[8]),
                    "ut": _num(row[9]) if len(row) > 9 else None,
                    "round": row[1],
                    "date": str(row[4] or "")[:10],
                    "shared_cell": len(toks) > 1,   # 복수 성적서 1행 → VT 분배 불명
                }
    wb.close()
    return ledger


def load_ceb_unit_index(xlsm_path: Path) -> dict:
    """CEB 청구 → 부재 단가 색인 + 집계.

    Returns:
      {
        "unit": {(building_norm, product_norm): {vt_unit, dd_qty, vt_total, drawing, rows}},
        "vt_total_all": float,
      }
    """
    import openpyxl

    unit: dict[tuple[str, str], dict] = {}
    vt_total_all = 0.0
    wb = openpyxl.load_workbook(str(xlsm_path), read_only=True, data_only=True)
    ws = wb["CEB"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not (row[1] or row[2]):
            continue
        b = normalize_building(row[2])
        p = normalize_product(row[5])
        vt_tot = _num(row[11]) or 0
        vt_total_all += vt_tot
        if not b or not p:
            continue
        key = (b, p)
        e = unit.setdefault(key, {"vt_unit": None, "dd_qty": 0.0, "vt_total": 0.0,
                                   "drawing": str(row[23] or ""), "rows": 0})
        e["rows"] += 1
        e["dd_qty"] += _num(row[9]) or 0
        e["vt_total"] += vt_tot
        vu = _num(row[10])
        if vu:
            # 동일 (건물,부재) 에 unit 상이하면 보수적으로 미확정 처리
            if e["vt_unit"] is None:
                e["vt_unit"] = vu
            elif e["vt_unit"] != vu:
                e["vt_unit"] = "CONFLICT"
    wb.close()
    return {"unit": unit, "vt_total_all": vt_total_all}


def _num(v):
    try:
        f = float(v)
        return f
    except (TypeError, ValueError):
        return None


# ─────────────────────────── 교차검증 엔진 ───────────────────────────


@dataclass
class CrosscheckResult:
    n_reports: int = 0
    n_nis: int = 0
    joined: list[dict] = field(default_factory=list)          # 물량표 조인 성공
    ledger_missing: list[dict] = field(default_factory=list)  # 성적서 有, 물량표 無 (미지급 후보)
    qty_match: list[dict] = field(default_factory=list)       # 수량 대사 일치
    qty_mismatch: list[dict] = field(default_factory=list)    # 수량 대사 불일치 (⚠ 과다 방향 검토)
    qty_incalculable: list[dict] = field(default_factory=list)  # unit Q'ty 미확보 등
    piece_overrun: list[dict] = field(default_factory=list)   # 검사부재수 > 도면 DD (초과)
    coverage_note: str = ""

    def to_dict(self) -> dict:
        return {k: getattr(self, k) for k in (
            "n_reports", "n_nis", "joined", "ledger_missing", "qty_match",
            "qty_mismatch", "qty_incalculable", "piece_overrun", "coverage_note")}


def crosscheck(
    reports: list[CivilReport],
    ledger: dict[str, dict],
    ceb: dict,
    *,
    tolerance_pct: float = 1.0,
) -> CrosscheckResult:
    res = CrosscheckResult(n_reports=len(reports))
    unit_idx = ceb["unit"]
    pieces_by_key: dict[tuple[str, str], float] = defaultdict(float)

    for r in reports:
        if r.org != "NIS":
            continue
        res.n_nis += 1
        led = ledger.get(r.report_no)
        bldg = normalize_building(r.facility) or normalize_building(
            building_from_drawing(r.drawing))

        base = {"report_no": r.report_no, "date": r.date, "building": bldg,
                "items": r.items, "page": r.start_page}

        if led is None:
            res.ledger_missing.append({**base,
                "note": "물량표(NIS 지급)에 없음 — 미지급 후보 또는 물량표 기준시점 이후. 지급의무 룰 참고"})
            continue
        res.joined.append({**base, "ledger_vt": led["vt"], "ledger_sheet": led["sheet"]})

        # ② 수량 대사: Σ(부재수 × unit Q'ty) vs 물량표 VT
        expected = 0.0
        unknown: list[str] = []
        for prod, pieces in r.items:
            e = unit_idx.get((bldg, prod)) if bldg else None
            if not e or e["vt_unit"] in (None, "CONFLICT"):
                unknown.append(prod)
                continue
            expected += pieces * e["vt_unit"]
            pieces_by_key[(bldg, prod)] += pieces

        led_vt = led["vt"] or 0
        if unknown or not r.items:
            res.qty_incalculable.append({**base, "ledger_vt": led_vt,
                "unknown_products": unknown,
                "note": ("성적서 부재 미추출" if not r.items else
                         f"CEB 에서 unit Q'ty 미확보: {unknown}")})
        elif led["shared_cell"]:
            res.qty_incalculable.append({**base, "ledger_vt": led_vt,
                "expected_points": expected,
                "note": "물량표 1행에 복수 성적서 — VT 분배 불명, 묶음 단위 대사 필요"})
        else:
            diff_pct = abs(expected - led_vt) / led_vt * 100 if led_vt else (
                0 if expected == 0 else 100)
            rec = {**base, "expected_points": expected, "ledger_vt": led_vt,
                   "diff_pct": round(diff_pct, 1)}
            (res.qty_match if diff_pct <= tolerance_pct else res.qty_mismatch).append(rec)

    # ③ 부재수 대사 — 검사부재 vs 도면 DD (초과만 표시: 과다 방향)
    for key, pieces in pieces_by_key.items():
        e = unit_idx.get(key)
        if e and e["dd_qty"] and pieces > e["dd_qty"]:
            res.piece_overrun.append({
                "building": key[0], "product": key[1],
                "inspected_pieces": pieces, "dd_qty": e["dd_qty"],
                "note": "성적서 부재수 > 도면 수량 — 설계변경 또는 중복 성적서 확인"})

    n_led = len(ledger)
    pct = f"{res.n_nis / n_led * 100:.0f}%" if n_led else "—"
    res.coverage_note = (
        f"보유 성적서 기준 커버리지: NIS 성적서 {res.n_nis}건 / 물량표 {n_led}건 "
        f"({pct}). '성적서 미보유' 는 발견사항 아님 (표본 한계)."
    )
    return res


# ─────────────────────────── 실행 헬퍼 ───────────────────────────


def run_crosscheck(
    reports_pdf: Path, ledger_xlsx: Path, claims_xlsm: Path,
    *, out_json: Optional[Path] = None,
) -> CrosscheckResult:
    """전체 실행: PDF 추출(캐시) → 파싱 → 3자 조인."""
    from app.extractors.civil_report_parser import extract_reports
    from app.extractors.pdf_extractor import extract

    extracted = extract(Path(reports_pdf), force_ocr=True)   # 디스크 캐시 hit 시 즉시
    pages = [p.text for p in extracted.pages]
    reports = extract_reports(pages)
    ledger = load_nis_ledger(Path(ledger_xlsx))
    ceb = load_ceb_unit_index(Path(claims_xlsm))
    res = crosscheck(reports, ledger, ceb)

    if out_json:
        out_json.parent.mkdir(parents=True, exist_ok=True)
        out_json.write_text(
            json.dumps(res.to_dict(), ensure_ascii=False, indent=2, default=str),
            encoding="utf-8")
        logger.info("교차검증 결과 저장: %s", out_json)
    return res
