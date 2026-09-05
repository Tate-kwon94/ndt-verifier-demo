"""청구 엑셀 파서 — 헤더 자동 탐지 + templates.yaml 매핑.

전제: 시공사이 발행하는 엑셀 양식이 고정되어 있지 않음 → 헤더 후보 매칭으로 유연하게 처리.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Optional

import openpyxl

from app.config import templates_config

logger = logging.getLogger(__name__)


@dataclass
class ParsedBilling:
    discipline: str
    sheet_name: str
    header_row: int                  # 1-based
    rows: list[dict]
    warnings: list[str]
    skipped_invalid_rows: list[dict] = None  # joint_no/ndt_method 없는 footer/주석 행

    def __post_init__(self):
        if self.skipped_invalid_rows is None:
            self.skipped_invalid_rows = []


@dataclass
class PerMethodEnrichment:
    """분야별 시트에서 추출한 (report_no → 도면번호·WPS) 매핑.

    Meridian CP-P1 엑셀의 Total 시트에는 Detailed Drawing 컬럼이 없으나
    PT/UT/VT(UNITx) 시트에는 있음. Total 행에 join 으로 보강.
    """
    by_report_no: dict[str, dict]   # report_no → {'drawing_no': ..., 'welding_map': ...}
    source_sheets: list[str]
    warnings: list[str]


# ─────────────────────────── Helpers ───────────────────────────


def _normalize_token(s) -> str:
    return re.sub(r"\s+", "", str(s or "")).strip().lower()


def _build_method_map() -> dict[str, str]:
    cfg = templates_config().get("ndt_method_normalization", {})
    m: dict[str, str] = {}
    for canonical, aliases in cfg.items():
        for a in aliases:
            m[_normalize_token(a)] = canonical
    return m


def _build_result_map() -> dict[str, str]:
    cfg = templates_config().get("result_normalization", {})
    m: dict[str, str] = {}
    for canonical, aliases in cfg.items():
        for a in aliases:
            m[_normalize_token(a)] = canonical
    return m


def _to_date(value) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _to_float(value) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return None


# ─────────────────────────── Discipline detection ───────────────────────────


def _detect_discipline(sheet_name: str, header_cells: Iterable[str]) -> Optional[str]:
    cfg = templates_config()["disciplines"]
    name_l = sheet_name.lower()
    for code, spec in cfg.items():
        for pattern in spec.get("sheet_name_patterns", []):
            if pattern.lower() in name_l:
                return code
    # 헤더에 분야 명시 시
    joined = " ".join(str(c or "") for c in header_cells).lower()
    for code in cfg:
        if code.lower() in joined:
            return code
    return None


# ─────────────────────────── Header row detection ───────────────────────────


def _find_header_row_and_cells(
    ws, search_max: int, candidate_tokens: set[str]
) -> tuple[Optional[int], list[str]]:
    """앞쪽 search_max 행만 읽어서 후보 토큰 가장 많이 매칭되는 행을 헤더로 채택.

    read_only 모드에서 ws.cell() 은 매우 느리므로 iter_rows 로 순차 읽기.
    헤더 셀(정규화 토큰 리스트)도 함께 반환해 후속 column_map 에서 재사용.
    """
    threshold = max(2, len(candidate_tokens) // 4)
    best_row, best_count, best_cells = None, 0, []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=search_max, values_only=True), start=1):
        cells = [_normalize_token(v) for v in row]
        count = sum(1 for c in cells if c in candidate_tokens)
        if count > best_count:
            best_row, best_count, best_cells = row_idx, count, cells
    if best_count < threshold:
        return None, []
    return best_row, best_cells


def _column_map_from_cells(
    header_cells: list[str], columns_spec: list[dict]
) -> tuple[dict[str, int], list[str]]:
    """field → column index (1-based) 매핑. 누락 필드 경고 반환."""
    mapping: dict[str, int] = {}
    warnings: list[str] = []
    for spec in columns_spec:
        field = spec["field"]
        for cand in spec["candidates"]:
            tok = _normalize_token(cand)
            if tok in header_cells:
                mapping[field] = header_cells.index(tok) + 1
                break
        else:
            if spec.get("required"):
                warnings.append(f"필수 컬럼 누락: {field} (후보: {spec['candidates']})")
    return mapping, warnings


# ─────────────────────────── Public ───────────────────────────


def parse_billing_xlsx(path: Path, *, discipline_hint: Optional[str] = None) -> ParsedBilling:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    cfg = templates_config()["disciplines"]
    candidate_tokens: set[str] = set()
    for code in (cfg if not discipline_hint else [discipline_hint]):
        for col in (cfg.get(code) or {}).get("columns", []):
            for c in col.get("candidates", []):
                candidate_tokens.add(_normalize_token(c))

    best: Optional[ParsedBilling] = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 시트명만으로 분야 탐지 — 매칭 안 되면 즉시 skip (read_only 모드에서 셀 접근 회피)
        discipline = discipline_hint or _detect_discipline_by_name(sheet_name)
        if discipline is None:
            continue
        spec = cfg.get(discipline)
        if not spec or not spec.get("columns"):
            continue

        search_max = spec.get("header_row_search_max", 10)
        header_row, header_cells = _find_header_row_and_cells(ws, search_max, candidate_tokens)
        if header_row is None:
            continue

        mapping, warnings = _column_map_from_cells(header_cells, spec["columns"])
        rows: list[dict] = []
        skipped: list[dict] = []
        for r in _read_rows_streaming(ws, header_row, mapping):
            # 필수 키 (joint_no, ndt_method) 누락은 footer/주석 행으로 간주하고 skip
            # — 검토자에게 카운트만 보고 (silent 누락 금지)
            if r["raw_json"].get("_invalid"):
                skipped.append(r["raw_json"])
            else:
                rows.append(r)
        result = ParsedBilling(
            discipline=discipline,
            sheet_name=sheet_name,
            header_row=header_row,
            rows=rows,
            warnings=warnings,
            skipped_invalid_rows=skipped,
        )
        if skipped:
            result.warnings.append(
                f"필수 키(joint_no/ndt_method) 누락 행 {len(skipped)}건 skip "
                f"(footer/주석 가능성). 첫 행 예: {skipped[0]}"
            )
        # 가장 많은 행을 가진 시트 채택 (CP-P1 의 경우 'CP-P1(Total)')
        if best is None or len(result.rows) > len(best.rows):
            best = result

    if best is None:
        raise ValueError(
            f"엑셀에서 분야/헤더를 인식하지 못했습니다: {path}. "
            f"templates.yaml 의 candidates / sheet_name_patterns 를 확장하세요."
        )
    return best


def _detect_discipline_by_name(sheet_name: str) -> Optional[str]:
    cfg = templates_config()["disciplines"]
    name_l = sheet_name.lower()
    for code, spec in cfg.items():
        for pattern in spec.get("sheet_name_patterns", []):
            if pattern.lower() in name_l:
                return code
    return None


def _read_rows_streaming(ws, header_row: int, mapping: dict[str, int]):
    """iter_rows 로 순차 읽기 — read_only 모드에 최적화. 0-based 인덱스 주의.

    각 행의 raw_json 에 `_excel_row` (1-based 원본 엑셀 행번호) 를 저장해
    writer 가 정확한 위치에 검토 컬럼을 쓰도록 함.
    """
    method_map = _build_method_map()
    result_map = _build_result_map()
    # mapping 의 1-based 컬럼 인덱스를 0-based 로 변환해 row tuple 인덱싱
    col_offset = {f: c - 1 for f, c in mapping.items()}

    for offset, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True)):
        excel_row = header_row + 1 + offset
        raw = {f: (row[idx] if idx < len(row) else None) for f, idx in col_offset.items()}
        if all(v is None or str(v).strip() == "" for v in raw.values()):
            continue   # 빈 행 skip

        raw_json = {k: _serialize(v) for k, v in raw.items()}
        raw_json["_excel_row"] = excel_row
        normalized = {
            "billing_no": _str_or_none(raw.get("billing_no")),
            "report_no": _normalize_report_no(raw.get("report_no")),
            "joint_no": _str_or_none(raw.get("joint_no")),
            "line_no": _str_or_none(raw.get("line_no")),
            "welder_id": _str_or_none(raw.get("welder_id")),
            "ndt_method": method_map.get(_normalize_token(raw.get("ndt_method")), _str_or_none(raw.get("ndt_method"))),
            "drawing_no": _str_or_none(raw.get("drawing_no")),
            "inspection_date": _to_date(raw.get("inspection_date")),
            "result": result_map.get(_normalize_token(raw.get("result")), _str_or_none(raw.get("result"))),
            "unit": _str_or_none(raw.get("unit")),
            "bldg": _str_or_none(raw.get("bldg")),
            "dimension": _str_or_none(raw.get("dimension")),
            "quantity": _to_float(raw.get("quantity")),
            "unit_price": _to_float(raw.get("unit_price")),
            "amount": _to_float(raw.get("amount")),
            "raw_json": raw_json,
        }
        # 필수 키 (joint_no, ndt_method) 누락된 행은 표시 → 검토자 확인 대상
        if not normalized["joint_no"] or not normalized["ndt_method"]:
            normalized["raw_json"]["_invalid"] = True
        yield normalized


def _normalize_report_no(v) -> Optional[str]:
    """성적서번호 정규화 — 공백·하이픈 차이로 매칭 실패 막기 위해 표시는 원본 그대로 두되 매칭 시 별도 정규화."""
    s = _str_or_none(v)
    return s


# ─────────────────────────── Per-method enrichment ───────────────────────────


def enrich_from_per_method_sheets(path: Path, *, discipline: str) -> PerMethodEnrichment:
    """동일 청구 엑셀의 분야별 시트(VT/PT/UT)에서 report_no → (drawing_no, welding_map) 추출.

    Total 시트에 빠진 정보를 채우는 보조 데이터. 시트별로 헤더는 거의 동일하지만
    column 위치만 다름 (예: VT 는 c20·c21, PT/UT 는 c16·c17). 헤더 토큰으로 찾음.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    by_report_no: dict[str, dict] = {}
    source_sheets: list[str] = []
    warnings: list[str] = []

    # 분야별 시트는 시트명에 NDT 방법 약자가 들어 있음 (VT/PT/UT/MT/RT)
    method_re = re.compile(r"^(VT|PT|UT|MT|RT)\s*[\(\[]", re.IGNORECASE)

    for sheet_name in wb.sheetnames:
        if not method_re.match(sheet_name):
            continue
        ws = wb[sheet_name]
        # 헤더 찾기 — 첫 6행 중 'Report Number' 또는 'Detailed Drawing' 포함 행
        target = {_normalize_token("Report Number"), _normalize_token("Detailed Drawing"),
                  _normalize_token("Welding Map")}
        header_row = None
        header_cells: list[str] = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
            cells = [_normalize_token(v) for v in row]
            if sum(1 for c in cells if c in target) >= 2:
                header_row = r_idx
                header_cells = cells
                break
        if header_row is None:
            warnings.append(f"{sheet_name}: 헤더에 'Report Number'/'Detailed Drawing' 미발견 — skip")
            continue

        def _col(tok: str) -> Optional[int]:
            try:
                return header_cells.index(_normalize_token(tok)) + 1
            except ValueError:
                return None

        rn_col = _col("Report Number")
        dwg_col = _col("Detailed Drawing")
        wm_col = _col("Welding Map")
        if not rn_col or not dwg_col:
            warnings.append(f"{sheet_name}: report_no 또는 drawing 컬럼 위치 미확정 — skip")
            continue

        source_sheets.append(sheet_name)
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            rn = _str_or_none(row[rn_col - 1] if rn_col - 1 < len(row) else None)
            if not rn:
                continue
            dwg = _str_or_none(row[dwg_col - 1] if dwg_col - 1 < len(row) else None)
            wm = _str_or_none(row[wm_col - 1] if wm_col and wm_col - 1 < len(row) else None)
            if not dwg and not wm:
                continue
            existing = by_report_no.get(rn)
            if existing is None:
                by_report_no[rn] = {"drawing_no": dwg, "welding_map": wm}
            else:
                # 동일 report_no 가 여러 시트에 — 값 충돌 검증
                if dwg and existing.get("drawing_no") and existing["drawing_no"] != dwg:
                    warnings.append(
                        f"report_no={rn}: 시트 간 drawing_no 불일치 "
                        f"({existing['drawing_no']} vs {dwg}) — 최초 값 유지"
                    )
                if wm and existing.get("welding_map") and existing["welding_map"] != wm:
                    warnings.append(
                        f"report_no={rn}: 시트 간 welding_map 불일치 "
                        f"({existing['welding_map']} vs {wm}) — 최초 값 유지"
                    )

    return PerMethodEnrichment(
        by_report_no=by_report_no,
        source_sheets=source_sheets,
        warnings=warnings,
    )


def _str_or_none(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _serialize(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v
