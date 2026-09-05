"""검사기준 가이드 생성 — 발주처 검토자 핸드북.

도면(DC) + 적재된 표준(NP/PNAE/SCWEP) + 청구 데이터 를 종합해
Safety Class 매트릭스·KKS 코드별 NDT 요구사항·표준 인용 cheat sheet 를 자동 생성.

출력:
  data/outputs/검사기준_가이드_<discipline>_<date>.xlsx   (시트별 검색용)
  data/outputs/검사기준_가이드_<discipline>_<date>.md     (검토자 핸드북)

검토 회의·시공사 협의 자료 직접 활용 가능.
"""
from __future__ import annotations

import collections
import html
import json
import re
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl
import pdfplumber
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select

from app.config import DATA_DIR
from app.database.models import (
    BillingItem,
    BillingRound,
    DrawingFile,
    DrawingSet,
    Requirement,
    StandardDocument,
    get_session,
    init_db,
)


# 표준·코드 인용 정규식 — Code Gap Analysis 용
_CODE_PATTERNS: dict[str, re.Pattern[str]] = {
    "NP":      re.compile(r"\bNP-?\d{3}(?:-\d{2,4})\b"),
    "PNAE":    re.compile(r"\bPNAE[\s-]*[A-Z]?[\s-]*\d+(?:[\s-]*\d+)*", re.IGNORECASE),
    "GOST":    re.compile(r"\bGOST\s+(?:ISO\s+)?\d{2,5}[\-\.\d]*"),
    "TU":      re.compile(r"\bTU\s+\d{1,3}[-\.\s\d/]{3,}", re.IGNORECASE),
    "ASME":    re.compile(r"\bASME\s+(?:Sec(?:tion)?\.?[\s\d]*(?:Div\.?\s*\d+)?|B\d+(?:\.\d+)?|Code\s+Case)\s*[^\n,]{0,30}", re.IGNORECASE),
    "RCC-M":   re.compile(r"\bRCC-M[\s-]*[A-Z]?[\s\d\-\.]{0,20}", re.IGNORECASE),
    "ISO":     re.compile(r"\bISO\s+\d{2,5}[\-\d]*"),
    "EN":      re.compile(r"\bEN\s+\d{2,5}[\-\d]*"),
    "KS":      re.compile(r"\bKS\s+[A-Z]\s*\d{2,5}"),
}


# ─────────────────────────── Public ───────────────────────────


def generate(
    *,
    discipline: str,
    out_dir: Optional[Path] = None,
) -> dict:
    """검사기준 가이드 생성. 통계 dict 반환 (생성된 파일 경로 포함)."""
    init_db()
    out_dir = out_dir or (DATA_DIR / "outputs")
    out_dir.mkdir(parents=True, exist_ok=True)

    today = date.today().isoformat()
    xlsx_path = out_dir / f"검사기준_가이드_{discipline}_{today}.xlsx"
    md_path = out_dir / f"검사기준_가이드_{discipline}_{today}.md"
    html_path = out_dir / f"검사기준_가이드_{discipline}_{today}.html"

    with get_session() as s:
        snapshot = _collect_snapshot(s, discipline)

    _write_xlsx(snapshot, xlsx_path)
    _write_md(snapshot, md_path)
    _write_html(snapshot, html_path)

    return {
        "xlsx": str(xlsx_path),
        "md": str(md_path),
        "html": str(html_path),
        "drawing_sets": len(snapshot["drawing_sets"]),
        "kks_lines": len(snapshot["kks_safety"]),
        "inspection_rows": len(snapshot["inspection_matrix"]),
        "standards_indexed": len(snapshot["standards"]),
        "billing_drawings_referenced": len(snapshot["billing_drawings"]),
        "code_gap_cited_only": len(snapshot["code_gap"]["cited_only"]),
        "code_gap_matched": len(snapshot["code_gap"]["matched"]),
    }


# ─────────────────────────── Snapshot ───────────────────────────


def _collect_snapshot(session, discipline: str) -> dict:
    """DB 에서 가이드 생성에 필요한 모든 자료를 한꺼번에 모음."""
    drawing_sets = list(session.scalars(select(DrawingSet)))
    standards = list(session.scalars(select(StandardDocument)))

    kks_safety: list[dict] = []        # Safety Class 매트릭스 행
    inspection_matrix: list[dict] = []  # KKS code × NDT 요구사항
    drawing_codes: dict[str, set[str]] = collections.defaultdict(set)  # drawing_no → 인용 표준
    drawing_to_lines: dict[str, set[str]] = collections.defaultdict(set)

    for ds in drawing_sets:
        combined = ds.combined_json or {}
        dwg_no = ds.drawing_no

        # 1ULD.DC 형식 (kks_lines + inspection_matrix) 우선
        for line in combined.get("kks_lines", []) or []:
            line = dict(line)
            line["drawing_no"] = dwg_no
            kks_safety.append(line)
            if line.get("kks_code"):
                drawing_to_lines[dwg_no].add(line["kks_code"])
        for ins in combined.get("inspection_matrix", []) or []:
            ins = dict(ins)
            ins["drawing_no"] = dwg_no
            inspection_matrix.append(ins)
            if ins.get("kks_code"):
                drawing_to_lines[dwg_no].add(ins["kks_code"])

        # 표 추출 결과가 없거나 옛 스키마 (joints) 인 경우의 폴백
        for joint in combined.get("joints", []) or []:
            ins_row = {
                "drawing_no": dwg_no,
                "kks_code": joint.get("line_no") or joint.get("joint_no"),
                "joint_no": joint.get("joint_no"),
                "joint_category": None,
                "safety_class_np_001_15": joint.get("safety_class"),
                "weld_type": joint.get("weld_type"),
                "vt_pct": None,
                "pt_or_mt_pct": None,
                "rt_pct": None,
                "ut_pct": None,
            }
            for n in joint.get("required_ndt", []) or []:
                m = (n.get("method") or "").upper()
                pct = n.get("sampling_rate_pct")
                key_map = {"VT": "vt_pct", "PT": "pt_or_mt_pct", "MT": "pt_or_mt_pct",
                           "RT": "rt_pct", "UT": "ut_pct"}
                if m in key_map and pct is not None:
                    ins_row[key_map[m]] = pct
            inspection_matrix.append(ins_row)

        # 인용 코드
        for ac in combined.get("applicable_codes", []) or []:
            code = ac.get("code")
            if code:
                drawing_codes[dwg_no].add(code)
        for joint in combined.get("joints", []) or []:
            for ac in joint.get("applicable_codes", []) or []:
                code = ac.get("code")
                if code:
                    drawing_codes[dwg_no].add(code)

    # 청구 엑셀에서 참조된 도면 분포 (CP-P1 만)
    billing_drawings = collections.Counter()
    rounds_for_disc = [
        r.id for r in session.scalars(select(BillingRound).where(BillingRound.discipline == discipline))
    ]
    if rounds_for_disc:
        for bi in session.scalars(
            select(BillingItem).where(BillingItem.billing_round_id.in_(rounds_for_disc))
        ):
            if bi.drawing_no:
                billing_drawings[bi.drawing_no] += 1

    # Code Gap Analysis — 도면에서 인용된 코드 ↔ 적재된 표준 매핑
    code_gap = _build_code_gap(session, drawing_sets, standards, drawing_codes)

    # 과다/누락 청구 집계 — 발주처 검토 1순위 산출물
    overbill_underbill = _build_overbill_underbill(session, discipline, inspection_matrix)

    return {
        "discipline": discipline,
        "generated_at": date.today().isoformat(),
        "drawing_sets": drawing_sets,
        "standards": standards,
        "kks_safety": kks_safety,
        "inspection_matrix": inspection_matrix,
        "drawing_codes": drawing_codes,
        "drawing_to_lines": drawing_to_lines,
        "billing_drawings": billing_drawings,
        "code_gap": code_gap,
        "overbill_underbill": overbill_underbill,
    }


# ─────────────────────────── 과다/누락 청구 집계 ───────────────────────────


_OVERBILL_CONFIRMED_RULE = "billed_ndt_not_in_requirements"
_OVERBILL_UNPROVEN_RULES = {"billed_ndt_basis_not_submitted", "billed_ndt_basis_unclear",
                            "billed_ndt_covered_by_scwep"}


def _split_overbill_by_basis(session, overbill: list[dict]) -> tuple[list[dict], list[dict]]:
    """도면 차집합으로 잡힌 과다청구 후보를 Finding 의 근거 상태로 둘로 가른다."""
    from app.database.models import Finding
    ids = [o["billing_item_id"] for o in overbill if o.get("billing_item_id") is not None]
    by_item: dict[int, Finding] = {}
    if ids:
        for f in session.scalars(select(Finding).where(Finding.billing_item_id.in_(ids))):
            by_item[f.billing_item_id] = f
    confirmed, unproven = [], []
    for o in overbill:
        f = by_item.get(o.get("billing_item_id"))
        rules = {fe.get("rule") for fe in (((f.citations_json or {}).get("findings") if f else None) or [])}
        if _OVERBILL_CONFIRMED_RULE in rules:
            status, reason = "confirmed", "no_basis_found"
        elif rules & _OVERBILL_UNPROVEN_RULES:
            status = "unproven"
            fe = next(fe for fe in f.citations_json["findings"] if fe.get("rule") in _OVERBILL_UNPROVEN_RULES)
            reason = (fe.get("details") or {}).get("basis_state") or "?"
        else:
            status, reason = "unproven", "not_reviewed"
        row = {**o, "basis_status": status, "basis_reason": reason}
        (confirmed if status == "confirmed" else unproven).append(row)
    return confirmed, unproven


def _build_overbill_underbill(session, discipline: str, inspection_matrix: list[dict]) -> dict:
    """청구의 (drawing_no, joint_no) 단위로 도면 요구 NDT 와 청구 NDT 를 차집합 비교.

    출력:
      - overbill: 도면 미요구 NDT 가 청구된 행 목록 (발주처 환불 후보)
      - underbill_joints: 도면 요구 NDT 중 청구되지 않은 (joint, method) 목록
      - per_joint: (drawing, joint) 별 요구·청구·과다·누락 요약
      - summary: 전체 통계
    """
    # 1) 도면 요구 매트릭스 → (drawing_prefix, kks_code) → set of NDT methods
    #    kks_code 가 청구의 joint_no 와 직접 매칭은 어렵지만 drawing prefix 단위 요구는 알 수 있음
    drawing_required: dict[str, set[str]] = collections.defaultdict(set)   # drawing 본체 → 요구 NDT
    for ins in inspection_matrix:
        dwg = ins.get("drawing_no") or ""
        body = _drawing_body(dwg)
        if not body:
            continue
        for method, pct_key in [("VT", "vt_pct"), ("PT", "pt_or_mt_pct"),
                                 ("MT", "pt_or_mt_pct"), ("RT", "rt_pct"),
                                 ("UT", "ut_pct")]:
            v = ins.get(pct_key)
            if v is not None and v != "" and v != 0 and str(v).strip() not in ("-", "0", "0%"):
                drawing_required[body].add(method)

    # 2) 청구의 (drawing_no, joint_no) → set of NDT methods
    rounds = list(session.scalars(
        select(BillingRound).where(BillingRound.discipline == discipline)
    ))
    round_ids = [br.id for br in rounds]
    if not round_ids:
        return {
            "overbill": [], "underbill_joints": [], "per_joint": [],
            "overbill_unproven": [],
            "summary": {
                "overbill_rows": 0,
                "overbill_total_amount": 0,
                "overbill_confirmed_rows": 0, "overbill_confirmed_amount": 0,
                "overbill_unproven_rows": 0, "overbill_unproven_amount": 0,
                "underbill_pairs": 0,
                "joints_analyzed": 0,
                "joints_with_required_known": 0,
                "drawings_with_requirements": len(drawing_required),
            },
        }

    billing_grouped: dict[tuple, dict] = {}  # (drawing_body, joint_no) → {methods, rows[]}
    for bi in session.scalars(select(BillingItem).where(BillingItem.billing_round_id.in_(round_ids))):
        dwg_body = _drawing_body(bi.drawing_no or "")
        if not dwg_body or not bi.joint_no or not bi.ndt_method:
            continue
        key = (dwg_body, bi.joint_no)
        slot = billing_grouped.setdefault(key, {"methods": set(), "rows": []})
        slot["methods"].add(bi.ndt_method.upper())
        slot["rows"].append({
            "billing_item_id": bi.id, "ndt_method": bi.ndt_method,
            "report_no": bi.report_no, "amount": bi.amount,
            "inspection_date": bi.inspection_date.isoformat() if bi.inspection_date else None,
        })

    # 3) 차집합 비교
    overbill: list[dict] = []     # 도면 미요구 NDT 가 청구된 청구 행
    underbill: list[dict] = []    # 도면 요구 NDT 중 청구 안 된 (joint, method)
    per_joint: list[dict] = []

    for (dwg_body, joint_no), slot in billing_grouped.items():
        required = drawing_required.get(dwg_body, set())
        billed = slot["methods"]
        if not required:
            # 도면 요구사항 모르면 비교 불가 — per_joint 에만 정보성 기록
            per_joint.append({
                "drawing_body": dwg_body, "joint_no": joint_no,
                "required_ndt": [], "billed_ndt": sorted(billed),
                "overbill_methods": [], "underbill_methods": [],
                "status": "도면 요구사항 미확정 (도면 DB 누락)",
            })
            continue

        over_methods = sorted(billed - required)
        under_methods = sorted(required - billed)
        for row in slot["rows"]:
            if row["ndt_method"].upper() in over_methods:
                overbill.append({
                    "drawing_body": dwg_body, "joint_no": joint_no,
                    "billed_method": row["ndt_method"],
                    "report_no": row["report_no"], "amount": row["amount"],
                    "inspection_date": row["inspection_date"],
                    "required_ndt": sorted(required),
                    "billing_item_id": row["billing_item_id"],
                })
        for m in under_methods:
            underbill.append({
                "drawing_body": dwg_body, "joint_no": joint_no,
                "missing_method": m,
                "billed_methods": sorted(billed),
                "required_ndt": sorted(required),
            })
        status = "OK"
        if over_methods and under_methods: status = "과다+누락"
        elif over_methods: status = "과다"
        elif under_methods: status = "누락"
        per_joint.append({
            "drawing_body": dwg_body, "joint_no": joint_no,
            "required_ndt": sorted(required), "billed_ndt": sorted(billed),
            "overbill_methods": over_methods, "underbill_methods": under_methods,
            "status": status,
        })

    # ── 2026-09-05 SCWEP 근거 게이트: 확정 vs 근거 미확정 분리 ──
    # 이 함수는 판정 결과(Finding)를 읽지 않고 도면 매트릭스 차집합으로 **독립적으로** 과다청구를
    # 계산해 금액과 함께 찍는다. 검토 엑셀은 "근거 제출 요청" 이라 하는데 이 문서는 "환불 후보 N원"
    # 이라 하면, 협상 테이블에 잘못된 쪽을 들고 나간다. Finding 을 billing_item_id 로 조인해
    # 확정(billed_ndt_not_in_requirements)만 '과다 청구' 로, 나머지는 '근거 미제출·미확정' 으로 가른다.
    # Finding 이 없는 행(미검토)은 보수적으로 미확정 쪽.
    overbill_confirmed, overbill_unproven = _split_overbill_by_basis(session, overbill)

    def _amt(rows):
        return sum((o.get("amount") or 0) for o in rows if isinstance(o.get("amount"), (int, float)))
    confirmed_amount = _amt(overbill_confirmed)
    unproven_amount = _amt(overbill_unproven)

    return {
        # 기존 키 'overbill' / 'overbill_rows' / 'overbill_total_amount' 는 **확정분만** — HTML 템플릿·시트가 그대로 읽는다.
        "overbill": overbill_confirmed,
        "overbill_unproven": overbill_unproven,
        "underbill_joints": underbill,
        "per_joint": per_joint,
        "summary": {
            "overbill_rows": len(overbill_confirmed),
            "overbill_total_amount": confirmed_amount,
            "overbill_confirmed_rows": len(overbill_confirmed),
            "overbill_confirmed_amount": confirmed_amount,
            "overbill_unproven_rows": len(overbill_unproven),
            "overbill_unproven_amount": unproven_amount,
            "underbill_pairs": len(underbill),
            "joints_analyzed": len(billing_grouped),
            "joints_with_required_known": sum(
                1 for p in per_joint if p["status"] != "도면 요구사항 미확정 (도면 DB 누락)"
            ),
            "drawings_with_requirements": len(drawing_required),
        },
    }


def _drawing_body(dwg: str) -> str:
    """도면번호에서 type+seq+rev 제거하고 본체만 반환.
    예: 'MD.D.P000.1.0UMA&&PAB&&&.021.DC.0001.E_C02' → 'MD.D.P000.1.0UMA&&PAB&&&.021'
    """
    if not dwg:
        return ""
    for marker in (".DC.", ".SD.", ".BG.", ".KE."):
        if marker in dwg:
            return dwg.split(marker)[0]
    return dwg


# ─────────────────────────── Code Gap Analysis ───────────────────────────


def _normalize_code(s: str) -> str:
    """공백·점·하이픈·트레일링 언어접미사 차이를 흡수해 매칭에 사용.

    예:
      'GOST 14098-2014'  → 'GOST140982014'
      'gost 14098-2014_en' → 'GOST140982014'  (언어접미사 제거)
      'pnaeg-7-010-89_en' → 'PNAEG710089'
      'PNAE G-7-010-89'   → 'PNAEG710089'
    """
    text = str(s)
    # 트레일링 언어접미사 제거 (파일명 stem 에 _en/_ru/_kor 가 자주 붙음)
    text = re.sub(r"_(?:en|ru|kor|kr|eng|jp|cn)$", "", text, flags=re.IGNORECASE)
    return re.sub(r"[\s\.\-_]+", "", text).upper()


def _build_code_gap(session, drawing_sets, standards, drawing_codes: dict) -> dict:
    """도면에서 콜한 모든 코드 ↔ 적재된 표준 매핑.

    수집원:
      - drawing_sets.combined_json 의 applicable_codes (LLM 추출본)
      - drawing_files 의 원본 PDF 텍스트 (정규식 추출, native 만)
    """
    cited: dict[str, dict] = {}   # normalized → {family, raw, citing_drawings: set, count}

    def _add(family: str, raw: str, drawing_no: Optional[str]):
        key = _normalize_code(raw)
        entry = cited.setdefault(key, {
            "family": family,
            "raw": raw,
            "citing_drawings": set(),
            "count": 0,
        })
        entry["count"] += 1
        if drawing_no:
            entry["citing_drawings"].add(drawing_no)

    # 1) LLM 추출 결과
    for dwg_no, codes in drawing_codes.items():
        for code in codes:
            fam = _family_for(code) or "OTHER"
            _add(fam, code, dwg_no)

    # 2) 원본 PDF 정규식 (native 만, scan 은 OCR 불가)
    drawing_files = list(session.scalars(select(DrawingFile)))
    for df in drawing_files:
        path = Path(df.file_path)
        if not path.exists() or "_scan" in path.name:
            continue
        try:
            with pdfplumber.open(path) as pdf:
                text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        except Exception:
            continue
        for fam, pat in _CODE_PATTERNS.items():
            for m in pat.finditer(text):
                raw = " ".join(m.group(0).split())
                _add(fam, raw, df.drawing_no)

    # 3) 청구회차의 성적서 PDF 도 같이 스캔 — 합격기준 인용이 풍부 (PNAE 713회 등)
    scanned_reports: set[str] = set()
    for br in session.scalars(select(BillingRound)):
        if not br.reports_pdf_path or br.reports_pdf_path in scanned_reports:
            continue
        scanned_reports.add(br.reports_pdf_path)
        rp = Path(br.reports_pdf_path)
        if not rp.exists():
            continue
        try:
            with pdfplumber.open(rp) as pdf:
                # 1000+ 페이지 성적서일 수 있어 텍스트 합산 — 매번 시간 들지만 정확성 우선
                text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        except Exception:
            continue
        report_label = f"[REPORT] {rp.name}"
        for fam, pat in _CODE_PATTERNS.items():
            for m in pat.finditer(text):
                raw = " ".join(m.group(0).split())
                _add(fam, raw, report_label)

    # 적재된 표준
    ingested: dict[str, dict] = {}   # normalized → {family, raw, doc_id, doc_type, path}
    for std in standards:
        for candidate in filter(None, [std.document_no, Path(std.file_path).stem]):
            key = _normalize_code(candidate)
            ingested.setdefault(key, {
                "family": _family_for(candidate) or std.doc_type.upper(),
                "raw": candidate,
                "doc_id": std.id,
                "doc_type": std.doc_type,
                "path": std.file_path,
            })

    cited_keys = set(cited.keys())
    ingested_keys = set(ingested.keys())

    matched = sorted(cited_keys & ingested_keys)
    cited_only = sorted(cited_keys - ingested_keys)
    ingested_only = sorted(ingested_keys - cited_keys)

    return {
        "cited": cited,
        "ingested": ingested,
        "matched": matched,
        "cited_only": cited_only,
        "ingested_only": ingested_only,
    }


def _family_for(code: str) -> Optional[str]:
    for fam, pat in _CODE_PATTERNS.items():
        if pat.search(code):
            return fam
    return None


# ─────────────────────────── XLSX writer ───────────────────────────


_HEADER_FILL = PatternFill("solid", fgColor="1976D2")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_NOTE_FILL = PatternFill("solid", fgColor="FFF9C4")


def _write_xlsx(snap: dict, path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _sheet_overview(wb, snap)
    _sheet_overbill_underbill(wb, snap)   # 발주처 검토 1순위
    _sheet_safety_matrix(wb, snap)
    _sheet_inspection_matrix(wb, snap)
    _sheet_drawing_codes(wb, snap)
    _sheet_billing_coverage(wb, snap)
    _sheet_code_gap(wb, snap)
    _sheet_standards_catalog(wb, snap)

    wb.save(path)


def _sheet_overbill_underbill(wb, snap):
    ws = wb.create_sheet("과다·누락 청구")
    ou = snap["overbill_underbill"]
    s = ou["summary"]

    # 상단 KPI
    ws.append([f"분야: {snap['discipline']}",
               f"분석된 (도면×Joint): {s['joints_analyzed']}",
               f"도면 요구사항 확보 Joint: {s['joints_with_required_known']}",
               f"⚠ 과다 청구 확정 행: {s['overbill_rows']}",
               f"근거 미제출·미확정 행: {s.get('overbill_unproven_rows', 0)}",
               f"⚠ 누락 (도면×NDT) 쌍: {s['underbill_pairs']}",
               f"과다 청구 확정 금액 합: {s.get('overbill_total_amount', 0):,.0f}",
               f"근거 미확정 금액 합 (환불 단정 아님): {s.get('overbill_unproven_amount', 0):,.0f}"])
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.append([])

    # 과다 청구 (1순위)
    ws.append(["⚠ 과다 청구 (도면이 요구하지 않은 NDT 가 청구됨 — 환불 후보)"])
    ws[ws.max_row][0].font = Font(bold=True, color="C62828")
    over_start = ws.max_row + 1
    _add_header(ws, over_start, [
        "도면 본체", "Joint No.", "청구 NDT (과다)",
        "성적서번호", "금액", "검사일", "도면 요구 NDT", "billing_item_id",
    ])
    for r_idx, o in enumerate(ou["overbill"], start=over_start + 1):
        ws.cell(row=r_idx, column=1, value=o["drawing_body"])
        ws.cell(row=r_idx, column=2, value=o["joint_no"])
        ws.cell(row=r_idx, column=3, value=o["billed_method"]).fill = PatternFill("solid", fgColor="FFCDD2")
        ws.cell(row=r_idx, column=4, value=o.get("report_no"))
        ws.cell(row=r_idx, column=5, value=o.get("amount"))
        ws.cell(row=r_idx, column=6, value=o.get("inspection_date"))
        ws.cell(row=r_idx, column=7, value=", ".join(o["required_ndt"]))
        ws.cell(row=r_idx, column=8, value=o["billing_item_id"])

    # 근거 미제출·미확정 (환불 단정 아님) — 2026-09-05 근거 게이트
    ws.append([])
    ws.append([f"근거 미제출·미확정 — 시공사에 근거 절차서(SCWEP) 제출 요청 (환불 단정 아님) "
               f"· {s.get('overbill_unproven_rows', 0)}행 · 금액 {s.get('overbill_unproven_amount', 0):,.0f}"])
    ws[ws.max_row][0].font = Font(bold=True, color="1565C0")
    unp_start = ws.max_row + 1
    _add_header(ws, unp_start, [
        "도면 본체", "Joint No.", "청구 NDT (도면 밖)", "근거 상태",
        "성적서번호", "금액", "검사일", "도면 요구 NDT", "billing_item_id",
    ])
    for r_idx, o in enumerate(ou.get("overbill_unproven", []), start=unp_start + 1):
        ws.cell(row=r_idx, column=1, value=o["drawing_body"])
        ws.cell(row=r_idx, column=2, value=o["joint_no"])
        ws.cell(row=r_idx, column=3, value=o["billed_method"]).fill = PatternFill("solid", fgColor="BBDEFB")
        ws.cell(row=r_idx, column=4, value=o.get("basis_reason"))
        ws.cell(row=r_idx, column=5, value=o.get("report_no"))
        ws.cell(row=r_idx, column=6, value=o.get("amount"))
        ws.cell(row=r_idx, column=7, value=o.get("inspection_date"))
        ws.cell(row=r_idx, column=8, value=", ".join(o["required_ndt"]))
        ws.cell(row=r_idx, column=9, value=o["billing_item_id"])

    # 누락 (2순위)
    ws.append([])
    ws.append(["⚠ 누락 (도면 요구 NDT 가 어느 청구 행에도 없음 — 시공 부적합 의심)"])
    ws[ws.max_row][0].font = Font(bold=True, color="EF6C00")
    under_start = ws.max_row + 1
    _add_header(ws, under_start, [
        "도면 본체", "Joint No.", "누락 NDT", "청구된 NDT", "도면 요구 NDT",
    ])
    for r_idx, u in enumerate(ou["underbill_joints"], start=under_start + 1):
        ws.cell(row=r_idx, column=1, value=u["drawing_body"])
        ws.cell(row=r_idx, column=2, value=u["joint_no"])
        ws.cell(row=r_idx, column=3, value=u["missing_method"]).fill = PatternFill("solid", fgColor="FFE0B2")
        ws.cell(row=r_idx, column=4, value=", ".join(u["billed_methods"]))
        ws.cell(row=r_idx, column=5, value=", ".join(u["required_ndt"]))

    # per_joint 상세
    ws.append([])
    ws.append(["전체 (Joint × NDT) 비교 — 도면 미확정 포함"])
    ws[ws.max_row][0].font = Font(bold=True)
    pj_start = ws.max_row + 1
    _add_header(ws, pj_start, [
        "도면 본체", "Joint No.", "도면 요구 NDT", "청구 NDT", "과다", "누락", "상태",
    ])
    for r_idx, p in enumerate(ou["per_joint"], start=pj_start + 1):
        ws.cell(row=r_idx, column=1, value=p["drawing_body"])
        ws.cell(row=r_idx, column=2, value=p["joint_no"])
        ws.cell(row=r_idx, column=3, value=", ".join(p["required_ndt"]))
        ws.cell(row=r_idx, column=4, value=", ".join(p["billed_ndt"]))
        ws.cell(row=r_idx, column=5, value=", ".join(p["overbill_methods"]))
        ws.cell(row=r_idx, column=6, value=", ".join(p["underbill_methods"]))
        cell = ws.cell(row=r_idx, column=7, value=p["status"])
        if "과다" in p["status"]: cell.fill = PatternFill("solid", fgColor="FFCDD2")
        elif "누락" in p["status"]: cell.fill = PatternFill("solid", fgColor="FFE0B2")
        elif "미확정" in p["status"]: cell.fill = PatternFill("solid", fgColor="ECEFF1")
        else: cell.fill = PatternFill("solid", fgColor="C8E6C9")


def _sheet_code_gap(wb, snap):
    ws = wb.create_sheet("코드 갭 분석")
    cited = snap["code_gap"]["cited"]
    ingested = snap["code_gap"]["ingested"]
    cited_only = snap["code_gap"]["cited_only"]
    matched = snap["code_gap"]["matched"]
    ingested_only = snap["code_gap"]["ingested_only"]

    ws.append([f"도면 인용 코드: {len(cited)}개", f"적재된 표준: {len(ingested)}개",
               f"매칭: {len(matched)}개", f"수령 필요(cited only): {len(cited_only)}",
               f"잉여(ingested only): {len(ingested_only)}"])
    ws.append([])

    _add_header(ws, 3, ["상태", "Family", "Code (raw)", "인용 횟수", "인용 도면 수", "인용 도면 (최대 3개)"])
    row = 4

    def write(status: str, fill: PatternFill, keys: list[str]):
        nonlocal row
        for key in keys:
            entry = cited.get(key) or ingested.get(key) or {}
            family = entry.get("family", "-")
            raw = entry.get("raw", key)
            count = entry.get("count", 0)
            citers = sorted(entry.get("citing_drawings", []))
            ws.cell(row=row, column=1, value=status).fill = fill
            ws.cell(row=row, column=2, value=family)
            ws.cell(row=row, column=3, value=raw)
            ws.cell(row=row, column=4, value=count)
            ws.cell(row=row, column=5, value=len(citers))
            ws.cell(row=row, column=6, value=", ".join(citers[:3]) + (f" ... +{len(citers)-3}" if len(citers) > 3 else ""))
            row += 1

    write("⚠ 수령 필요 (cited only)", PatternFill("solid", fgColor="FFCCBC"), cited_only)
    write("✓ 매칭 (in DB)",          PatternFill("solid", fgColor="C8E6C9"), matched)
    write("ℹ 잉여 (ingested only)",  PatternFill("solid", fgColor="E1F5FE"), ingested_only)


def _add_header(ws, row: int, headers: list[str]):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _sheet_overview(wb, snap):
    ws = wb.create_sheet("개요")
    ws.append(["NDT Assistant — 검사기준 가이드"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["분야", snap["discipline"]])
    ws.append(["생성일", snap["generated_at"]])
    ws.append(["도면 세트", len(snap["drawing_sets"])])
    ws.append(["KKS code (Safety Class 표 행)", len(snap["kks_safety"])])
    ws.append(["검사 매트릭스 행", len(snap["inspection_matrix"])])
    ws.append(["적재된 표준 문서", len(snap["standards"])])
    ws.append(["청구에 등장하는 고유 도면", len(snap["billing_drawings"])])
    ws.append([])
    ws.append(["사용법"])
    ws["A10"].font = Font(bold=True)
    ws.append(["1.", "Safety Class 매트릭스: KKS code 별 안전등급·seismic·QA 카테고리"])
    ws.append(["2.", "검사 요구 매트릭스: KKS code × NDT 방법(VT/PT-MT/RT/UT) 샘플링률"])
    ws.append(["3.", "도면-표준 인용 매핑: 어느 도면이 어떤 NP/PNAE/GOST 인용했는지"])
    ws.append(["4.", "청구 도면 커버리지: 청구 엑셀이 참조한 도면 중 가이드 DB 에 있는 비율"])
    ws.append(["5.", "표준 카탈로그: 적재된 NP/PNAE/SCWEP 문서 목록"])


def _sheet_safety_matrix(wb, snap):
    ws = wb.create_sheet("Safety Class 매트릭스")
    cols = [
        "도면번호", "KKS code", "재질", "Dout×t (mm)",
        "Safety class\n(NP-001-15)", "Classification\n(NP-001)",
        "Seismic\n(NP-031-01)", "QA category",
        "Design Pressure\n(MPag)", "Design Temp (°C)",
        "Operating P (MPag)", "Operating T (°C)",
        "Hydraulic Test P (MPag)", "Min Wall T (°C)",
        "Insulation", "page", "근거 (원문 인용)",
    ]
    _add_header(ws, 1, cols)
    for r_idx, line in enumerate(snap["kks_safety"], start=2):
        ws.cell(row=r_idx, column=1, value=line.get("drawing_no"))
        ws.cell(row=r_idx, column=2, value=line.get("kks_code"))
        ws.cell(row=r_idx, column=3, value=line.get("material"))
        ws.cell(row=r_idx, column=4, value=line.get("dout_x_thickness_mm"))
        ws.cell(row=r_idx, column=5, value=line.get("safety_class_np_001_15"))
        ws.cell(row=r_idx, column=6, value=line.get("classification_np_001"))
        ws.cell(row=r_idx, column=7, value=line.get("seismic_category_np_031_01"))
        ws.cell(row=r_idx, column=8, value=line.get("qa_category"))
        ws.cell(row=r_idx, column=9, value=line.get("design_pressure_mpag"))
        ws.cell(row=r_idx, column=10, value=line.get("design_temperature_c"))
        ws.cell(row=r_idx, column=11, value=line.get("operating_pressure_mpag"))
        ws.cell(row=r_idx, column=12, value=line.get("operating_temperature_c"))
        ws.cell(row=r_idx, column=13, value=line.get("hydraulic_test_pressure_mpag"))
        ws.cell(row=r_idx, column=14, value=line.get("min_wall_temperature_c"))
        ws.cell(row=r_idx, column=15, value=line.get("thermal_insulation"))
        ws.cell(row=r_idx, column=16, value=line.get("page_in_drawing"))
        ws.cell(row=r_idx, column=17, value=line.get("citation_quote"))


def _sheet_inspection_matrix(wb, snap):
    ws = wb.create_sheet("검사 요구 매트릭스")
    cols = [
        "도면번호", "KKS code", "Dout×t (mm)", "Joint Category",
        "Safety Class",
        "VT %", "PT/MT %", "RT %", "UT %",
        "보조용접 VT %", "보조용접 PT/MT %",
        "page", "근거 (원문 인용)",
    ]
    _add_header(ws, 1, cols)
    for r_idx, ins in enumerate(snap["inspection_matrix"], start=2):
        ws.cell(row=r_idx, column=1, value=ins.get("drawing_no"))
        ws.cell(row=r_idx, column=2, value=ins.get("kks_code"))
        ws.cell(row=r_idx, column=3, value=ins.get("dout_x_thickness_mm"))
        ws.cell(row=r_idx, column=4, value=ins.get("joint_category"))
        ws.cell(row=r_idx, column=5, value=ins.get("safety_class_np_001_15"))
        ws.cell(row=r_idx, column=6, value=ins.get("vt_pct"))
        ws.cell(row=r_idx, column=7, value=ins.get("pt_or_mt_pct"))
        ws.cell(row=r_idx, column=8, value=ins.get("rt_pct"))
        ws.cell(row=r_idx, column=9, value=ins.get("ut_pct"))
        ws.cell(row=r_idx, column=10, value=ins.get("auxiliary_vt_pct"))
        ws.cell(row=r_idx, column=11, value=ins.get("auxiliary_pt_or_mt_pct"))
        ws.cell(row=r_idx, column=12, value=ins.get("page_in_drawing"))
        ws.cell(row=r_idx, column=13, value=ins.get("citation_quote"))


def _sheet_drawing_codes(wb, snap):
    ws = wb.create_sheet("도면 ↔ 표준 인용")
    _add_header(ws, 1, ["도면번호", "인용한 표준 (콤마 구분)", "표준 수"])
    for r_idx, (dwg, codes) in enumerate(sorted(snap["drawing_codes"].items()), start=2):
        ws.cell(row=r_idx, column=1, value=dwg)
        ws.cell(row=r_idx, column=2, value=", ".join(sorted(codes)))
        ws.cell(row=r_idx, column=3, value=len(codes))


def _sheet_billing_coverage(wb, snap):
    ws = wb.create_sheet("청구 도면 커버리지")
    _add_header(ws, 1, ["청구의 도면번호", "참조 건수", "도면 DB 에 있음?", "비고 (KE=오기)"])
    indexed = {ds.drawing_no for ds in snap["drawing_sets"]}
    for r_idx, (dwg, n) in enumerate(snap["billing_drawings"].most_common(), start=2):
        # prefix 매칭 (DC/SD/BG/rev 제거 후 비교)
        body = dwg.split(".DC.")[0].split(".SD.")[0].split(".BG.")[0].split(".KE.")[0]
        has_in_db = any(idx.startswith(body) for idx in indexed)
        note = ""
        if ".KE." in dwg:
            note = "KE = WEP/SCWEP 오기 (시공사 정정 요청)"
        elif not has_in_db:
            note = "도면 DB 미적재 — 수령 필요"
        ws.cell(row=r_idx, column=1, value=dwg)
        ws.cell(row=r_idx, column=2, value=n)
        ws.cell(row=r_idx, column=3, value="✓" if has_in_db else "")
        cell = ws.cell(row=r_idx, column=4, value=note)
        if note:
            cell.fill = _NOTE_FILL


def _sheet_standards_catalog(wb, snap):
    ws = wb.create_sheet("표준 카탈로그")
    _add_header(ws, 1, ["문서 종류", "문서번호", "Rev.", "파일 경로", "적재일"])
    for r_idx, std in enumerate(sorted(snap["standards"], key=lambda s: (s.doc_type, s.file_path)), start=2):
        ws.cell(row=r_idx, column=1, value=std.doc_type)
        ws.cell(row=r_idx, column=2, value=std.document_no)
        ws.cell(row=r_idx, column=3, value=std.revision)
        ws.cell(row=r_idx, column=4, value=std.file_path)
        ws.cell(row=r_idx, column=5, value=std.ingested_at.isoformat() if std.ingested_at else None)


# ─────────────────────────── Markdown writer ───────────────────────────


def _write_html(snap: dict, path: Path) -> None:
    """단일 self-contained HTML — 인터넷 없이도 폐쇄망 PC 브라우저에서 열림.

    - 인라인 JSON 데이터 + vanilla JS (외부 라이브러리 없음)
    - 검색 / Safety Class·NDT 필터 / 도면 필터 / 표 정렬
    - 원본 인용은 그대로 (변형·요약 없음) — 검토자가 시공사·NIS 협의에 그대로 사용
    """
    payload = {
        "discipline": snap["discipline"],
        "generated_at": snap["generated_at"],
        "summary": {
            "drawing_sets": len(snap["drawing_sets"]),
            "kks_lines": len(snap["kks_safety"]),
            "inspection_rows": len(snap["inspection_matrix"]),
            "standards": len(snap["standards"]),
            "billing_drawings": len(snap["billing_drawings"]),
            "overbill_rows": snap["overbill_underbill"]["summary"]["overbill_rows"],
            "underbill_pairs": snap["overbill_underbill"]["summary"]["underbill_pairs"],
            "overbill_amount": snap["overbill_underbill"]["summary"].get("overbill_total_amount", 0),
            "overbill_unproven_rows": snap["overbill_underbill"]["summary"].get("overbill_unproven_rows", 0),
            "overbill_unproven_amount": snap["overbill_underbill"]["summary"].get("overbill_unproven_amount", 0),
        },
        "overbill_underbill": snap["overbill_underbill"],
        "safety_class_dist": dict(collections.Counter(
            (line.get("safety_class_np_001_15") or "(unspecified)") for line in snap["kks_safety"]
        )),
        "kks_safety": snap["kks_safety"],
        "inspection_matrix": snap["inspection_matrix"],
        "drawing_codes": {
            dwg: sorted(codes) for dwg, codes in snap["drawing_codes"].items()
        },
        "billing_drawings": [
            {"drawing_no": d, "count": n} for d, n in snap["billing_drawings"].most_common()
        ],
        "code_gap": {
            "cited_only": [
                {
                    "key": k,
                    "family": snap["code_gap"]["cited"][k]["family"],
                    "raw": snap["code_gap"]["cited"][k]["raw"],
                    "count": snap["code_gap"]["cited"][k]["count"],
                    "citing_drawings": sorted(snap["code_gap"]["cited"][k]["citing_drawings"]),
                } for k in snap["code_gap"]["cited_only"]
            ],
            "matched": [
                {
                    "key": k,
                    "family": snap["code_gap"]["cited"].get(k, {}).get("family")
                              or snap["code_gap"]["ingested"][k]["family"],
                    "raw": snap["code_gap"]["cited"].get(k, {}).get("raw")
                            or snap["code_gap"]["ingested"][k]["raw"],
                    "count": snap["code_gap"]["cited"].get(k, {}).get("count", 0),
                    "citing_drawings": sorted(snap["code_gap"]["cited"].get(k, {}).get("citing_drawings", [])),
                    "ingested_path": snap["code_gap"]["ingested"][k]["path"],
                } for k in snap["code_gap"]["matched"]
            ],
            "ingested_only": [
                {
                    "key": k,
                    "family": snap["code_gap"]["ingested"][k]["family"],
                    "raw": snap["code_gap"]["ingested"][k]["raw"],
                    "ingested_path": snap["code_gap"]["ingested"][k]["path"],
                } for k in snap["code_gap"]["ingested_only"]
            ],
        },
        "standards": [
            {
                "doc_type": std.doc_type,
                "document_no": std.document_no,
                "revision": std.revision,
                "file_path": std.file_path,
                "ingested_at": std.ingested_at.isoformat() if std.ingested_at else None,
            } for std in snap["standards"]
        ],
        "billing_indexed_prefixes": [ds.drawing_no for ds in snap["drawing_sets"]],
    }

    json_blob = json.dumps(payload, ensure_ascii=False, default=str)
    safe_title = html.escape(f"검사기준 가이드 — {snap['discipline']}")

    template = _HTML_TEMPLATE.replace("__TITLE__", safe_title)\
                              .replace("__GENERATED_AT__", html.escape(snap["generated_at"]))\
                              .replace("__DATA_JSON__", json_blob)
    path.write_text(template, encoding="utf-8")


_HTML_TEMPLATE = r"""<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<title>__TITLE__</title>
<style>
  :root { --primary:#1976d2; --bg:#fafafa; --border:#e0e0e0; --warn:#ff6f00; --ok:#2e7d32; --info:#0288d1; --muted:#757575; }
  * { box-sizing: border-box; }
  body { margin:0; font-family:-apple-system,'Helvetica Neue',Helvetica,'Apple SD Gothic Neo','Malgun Gothic',sans-serif; background:var(--bg); color:#212121; font-size:14px; line-height:1.5; }
  header { background:var(--primary); color:#fff; padding:16px 24px; }
  header h1 { margin:0; font-size:20px; }
  header .meta { opacity:0.85; font-size:12px; margin-top:4px; }
  nav { background:#fff; border-bottom:1px solid var(--border); padding:0 24px; position:sticky; top:0; z-index:10; overflow-x:auto; white-space:nowrap; }
  nav a { display:inline-block; padding:12px 16px; color:var(--muted); text-decoration:none; font-weight:500; border-bottom:3px solid transparent; }
  nav a.active, nav a:hover { color:var(--primary); border-bottom-color:var(--primary); }
  main { padding:24px; max-width:1600px; margin:0 auto; }
  section { background:#fff; padding:20px; border-radius:8px; box-shadow:0 1px 3px rgba(0,0,0,0.05); margin-bottom:24px; }
  section h2 { margin-top:0; color:var(--primary); font-size:18px; border-bottom:2px solid var(--primary); padding-bottom:8px; }
  .kpi-grid { display:grid; grid-template-columns:repeat(auto-fit,minmax(160px,1fr)); gap:12px; margin:16px 0; }
  .kpi { background:var(--bg); padding:12px 16px; border-radius:6px; border-left:4px solid var(--primary); }
  .kpi .label { color:var(--muted); font-size:12px; }
  .kpi .value { font-size:22px; font-weight:600; color:var(--primary); margin-top:4px; }
  .toolbar { display:flex; gap:8px; flex-wrap:wrap; margin-bottom:12px; align-items:center; }
  .toolbar input, .toolbar select { padding:6px 10px; border:1px solid var(--border); border-radius:4px; font-size:13px; }
  .toolbar input[type=search] { flex:1; min-width:200px; }
  .toolbar .chip { background:var(--bg); padding:4px 10px; border-radius:12px; font-size:12px; color:var(--muted); }
  table { width:100%; border-collapse:collapse; font-size:13px; }
  th, td { padding:8px 10px; text-align:left; border-bottom:1px solid var(--border); vertical-align:top; }
  th { background:var(--bg); cursor:pointer; user-select:none; white-space:nowrap; }
  th .sort-ind { color:var(--muted); margin-left:4px; }
  td.code, td.kks { font-family:'SF Mono','Monaco','Cascadia Mono',monospace; font-size:12px; }
  td.q { color:var(--ok); font-weight:600; text-align:right; }
  td.dash { color:var(--muted); text-align:right; }
  td.cite { color:#555; font-size:12px; max-width:480px; }
  .badge { display:inline-block; padding:2px 8px; border-radius:10px; font-size:11px; font-weight:600; }
  .badge.warn { background:#fff3e0; color:var(--warn); }
  .badge.ok { background:#e8f5e9; color:var(--ok); }
  .badge.info { background:#e1f5fe; color:var(--info); }
  .badge.ke { background:#fce4ec; color:#c2185b; }
  details { margin:8px 0; }
  details summary { cursor:pointer; padding:4px 0; color:var(--primary); }
  .note { background:#fff8e1; padding:10px 12px; border-left:4px solid var(--warn); border-radius:0 4px 4px 0; margin:12px 0; font-size:13px; }
  .small { color:var(--muted); font-size:12px; }
  tr.hidden { display:none; }
</style>
</head>
<body>
<header>
  <h1>__TITLE__</h1>
  <div class="meta">생성일: __GENERATED_AT__ · 발주처 기술검토자 핸드북 · 시공사·NIS 협의 근거 자료</div>
</header>
<nav id="topnav"></nav>
<main>

<section id="overview">
  <h2>1. 개요</h2>
  <div class="kpi-grid" id="kpi-grid"></div>
  <div class="note">이 가이드는 도면·표준·청구 데이터를 종합한 검토자 핸드북입니다. 모든 행은 도면 원문(표 그대로 추출)에 근거하며 추정·요약하지 않습니다. 검토 시 시공사·NIS 와 협의 자료로 그대로 인용 가능합니다.</div>
  <div>
    <strong>Safety Class 분포 (KKS code 수)</strong>
    <table id="sc-dist" style="max-width:400px"></table>
  </div>
</section>

<section id="overbill">
  <h2>2. ⚠ 과다·누락 청구 (발주처 검토 1순위)</h2>
  <div class="note">발주처 검토 핵심 = <strong>수량 검증</strong>. NDT 합격/불합격은 NIS 책임. 이 섹션은 (도면 본체 × Joint No.) 단위로 도면 요구 NDT 와 청구 NDT 를 차집합 비교한 결과입니다.</div>
  <div class="kpi-grid" id="ob-kpi"></div>
  <h3 style="margin-bottom:6px; color:#c62828">⚠ 과다 청구 (도면 미요구 NDT 가 청구됨 — 환불 후보)</h3>
  <div class="toolbar">
    <input type="search" id="ob-search" placeholder="도면 / Joint / NDT 검색...">
    <select id="ob-ndt-filter"><option value="">전체 NDT</option><option value="VT">VT</option><option value="PT">PT</option><option value="MT">MT</option><option value="RT">RT</option><option value="UT">UT</option></select>
    <span class="chip" id="ob-count"></span>
  </div>
  <div style="overflow-x:auto"><table id="ob-table"></table></div>

  <h3 style="margin-bottom:6px; color:#1565c0">근거 미제출·미확정 — 시공사에 근거 절차서(SCWEP) 제출 요청 (환불 단정 아님)</h3>
  <div class="note">도면에 없는 검사가 청구됐지만 <strong>근거 사슬이 아직 완전하지 않은</strong> 행입니다. 제출된 SCWEP 가 없거나(미제출), 적용 범위·조건이 확인되지 않았거나(불명확), SCWEP 조건부 요구가 이 검사를 지목합니다(근거 있음). 과다청구로 단정하지 않고 <strong>근거 제출을 요청</strong>하는 행이며, 금액도 위 확정 금액과 별도로 집계됩니다.</div>
  <div style="overflow-x:auto"><table id="ob-unproven"></table></div>

  <h3 style="margin-bottom:6px; color:#ef6c00">⚠ 누락 (도면 요구 NDT 가 어느 청구 행에도 없음)</h3>
  <div class="toolbar">
    <input type="search" id="ub-search" placeholder="도면 / Joint / NDT 검색...">
    <span class="chip" id="ub-count"></span>
  </div>
  <div style="overflow-x:auto"><table id="ub-table"></table></div>

  <details><summary>전체 (Joint × NDT) 비교 (도면 미확정 포함)</summary>
    <div class="toolbar">
      <input type="search" id="pj-search" placeholder="...">
      <select id="pj-status-filter">
        <option value="">전체 상태</option>
        <option value="과다">과다</option>
        <option value="누락">누락</option>
        <option value="과다+누락">과다+누락</option>
        <option value="OK">OK</option>
        <option value="미확정">도면 미확정</option>
      </select>
      <span class="chip" id="pj-count"></span>
    </div>
    <div style="overflow-x:auto"><table id="pj-table"></table></div>
  </details>
</section>

<section id="safety">
  <h2>3. Safety Class 매트릭스</h2>
  <div class="toolbar">
    <input type="search" id="safety-search" placeholder="KKS code / 도면 / 재질 검색...">
    <select id="safety-class-filter"><option value="">전체 Safety Class</option></select>
    <select id="safety-drawing-filter"><option value="">전체 도면</option></select>
    <span class="chip" id="safety-count"></span>
  </div>
  <div style="overflow-x:auto"><table id="safety-table"></table></div>
</section>

<section id="inspection">
  <h2>4. 검사 요구 매트릭스 (KKS code × NDT %)</h2>
  <div class="toolbar">
    <input type="search" id="ins-search" placeholder="KKS / 도면 / Joint 카테고리 검색...">
    <select id="ins-ndt-filter">
      <option value="">전체 NDT</option>
      <option value="vt_pct">VT 요구</option>
      <option value="pt_or_mt_pct">PT/MT 요구</option>
      <option value="rt_pct">RT 요구</option>
      <option value="ut_pct">UT 요구</option>
    </select>
    <span class="chip" id="ins-count"></span>
  </div>
  <div style="overflow-x:auto"><table id="ins-table"></table></div>
</section>

<section id="codegap">
  <h2>5. 코드 갭 분석 (도면 인용 ↔ 적재된 표준)</h2>
  <div class="kpi-grid" id="codegap-kpi"></div>
  <div class="note" id="codegap-note"></div>
  <h3 style="margin-bottom:8px">⚠ 수령 필요 (도면이 인용하나 표준 DB 에 없음)</h3>
  <div style="overflow-x:auto"><table id="codegap-cited-only"></table></div>
  <details><summary>✓ 매칭 (이미 적재)</summary>
    <div style="overflow-x:auto"><table id="codegap-matched"></table></div>
  </details>
  <details><summary>ℹ 잉여 (적재됐으나 도면에 안 보임)</summary>
    <div style="overflow-x:auto"><table id="codegap-ingested-only"></table></div>
  </details>
</section>

<section id="drawing-codes">
  <h2>6. 도면별 인용 표준</h2>
  <div class="toolbar"><input type="search" id="dc-search" placeholder="도면번호 / 표준 검색..."></div>
  <div style="overflow-x:auto"><table id="dc-table"></table></div>
</section>

<section id="billing">
  <h2>7. 청구 도면 커버리지</h2>
  <div class="toolbar">
    <input type="search" id="bd-search" placeholder="도면번호 검색...">
    <select id="bd-status-filter">
      <option value="">전체</option>
      <option value="missing">DB 미적재만</option>
      <option value="indexed">적재된 것만</option>
      <option value="ke">KE 오기만</option>
    </select>
    <span class="chip" id="bd-count"></span>
  </div>
  <div style="overflow-x:auto"><table id="bd-table"></table></div>
</section>

<section id="standards">
  <h2>8. 적재된 표준 카탈로그</h2>
  <div style="overflow-x:auto"><table id="std-table"></table></div>
</section>

</main>

<script id="payload" type="application/json">__DATA_JSON__</script>
<script>
"use strict";
const DATA = JSON.parse(document.getElementById('payload').textContent);

const SECTIONS = [
  ["overview","1. 개요"],["overbill","2. ⚠ 과다·누락 청구"],
  ["safety","3. Safety Class"],["inspection","4. 검사 매트릭스"],
  ["codegap","5. 코드 갭"],["drawing-codes","6. 도면-표준"],
  ["billing","7. 청구 커버리지"],["standards","8. 표준 카탈로그"],
];
function renderNav(){
  const n = document.getElementById('topnav');
  n.innerHTML = SECTIONS.map(([id,label])=>`<a href="#${id}">${label}</a>`).join('');
}

function renderKPIs(){
  const s = DATA.summary;
  const g = document.getElementById('kpi-grid');
  const items = [
    ["분야", DATA.discipline],
    ["⚠ 과다 청구 확정", (s.overbill_rows||0) + ' 행'],
    ["근거 미제출·미확정", (s.overbill_unproven_rows||0) + ' 행'],
    ["⚠ 누락 (Joint×NDT)", (s.underbill_pairs||0)],
    ["과다 확정 금액 합", (s.overbill_amount||0).toLocaleString()],
    ["근거 미확정 금액 (환불 단정 아님)", (s.overbill_unproven_amount||0).toLocaleString()],
    ["도면 세트", s.drawing_sets],
    ["적재된 표준", s.standards],
    ["청구 도면 (고유)", s.billing_drawings],
  ];
  g.innerHTML = items.map(([l,v])=>`<div class="kpi"><div class="label">${l}</div><div class="value">${v}</div></div>`).join('');

  const t = document.getElementById('sc-dist');
  const rows = Object.entries(DATA.safety_class_dist).sort();
  t.innerHTML = '<thead><tr><th>Safety Class</th><th style="text-align:right">KKS 수</th></tr></thead>' +
    '<tbody>' + rows.map(([k,v])=>`<tr><td>${esc(k)}</td><td style="text-align:right">${v}</td></tr>`).join('') + '</tbody>';
}

function esc(s){ if (s===null||s===undefined) return ''; return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }
function dashIfNull(v){ return (v===null||v===undefined||v==='') ? '<td class="dash">−</td>' : `<td class="q">${esc(v)}</td>`; }

function buildTable(elId, headers, rows, options){
  const opts = options || {};
  const el = document.getElementById(elId);
  const thead = '<thead><tr>' + headers.map((h,i)=>`<th data-col="${i}">${esc(h)}<span class="sort-ind"></span></th>`).join('') + '</tr></thead>';
  const tbody = '<tbody>' + rows.map(r=>'<tr>' + r.map((c,i)=> (opts.classes && opts.classes[i]) ? `<td class="${opts.classes[i]}">${c}</td>` : `<td>${c}</td>`).join('') + '</tr>').join('') + '</tbody>';
  el.innerHTML = thead + tbody;
  // sortable
  el.querySelectorAll('th').forEach(th=>{
    th.addEventListener('click', ()=>{
      const col = parseInt(th.dataset.col);
      const asc = th.dataset.dir !== 'asc';
      const trs = Array.from(el.tBodies[0].rows);
      trs.sort((a,b)=>{
        const av = a.cells[col].textContent.trim(); const bv = b.cells[col].textContent.trim();
        const an = parseFloat(av); const bn = parseFloat(bv);
        if (!isNaN(an) && !isNaN(bn)) return asc ? an-bn : bn-an;
        return asc ? av.localeCompare(bv) : bv.localeCompare(av);
      });
      el.tBodies[0].append(...trs);
      el.querySelectorAll('th').forEach(t=>{t.dataset.dir=''; t.querySelector('.sort-ind').textContent='';});
      th.dataset.dir = asc ? 'asc' : 'desc';
      th.querySelector('.sort-ind').textContent = asc ? '▲' : '▼';
    });
  });
}

function bindFilter(tableId, inputs, predicate, counterEl){
  const apply = ()=>{
    const tb = document.getElementById(tableId).tBodies[0];
    if (!tb) return;
    let visible = 0;
    Array.from(tb.rows).forEach(r=>{
      const show = predicate(r, inputs);
      r.classList.toggle('hidden', !show);
      if (show) visible++;
    });
    if (counterEl) counterEl.textContent = visible + ' / ' + tb.rows.length + ' 행';
  };
  Object.values(inputs).forEach(el=> el.addEventListener('input', apply));
  Object.values(inputs).forEach(el=> el.addEventListener('change', apply));
  apply();
}

// ─── Overbill / Underbill (발주처 검토 1순위) ───
function renderOverbill(){
  const ou = DATA.overbill_underbill;
  const s = ou.summary;
  document.getElementById('ob-kpi').innerHTML = [
    ["과다 청구 확정 행", s.overbill_rows],
    ["과다 청구 확정 금액 합", (s.overbill_total_amount||0).toLocaleString()],
    ["근거 미제출·미확정 행", s.overbill_unproven_rows||0],
    ["근거 미확정 금액 (환불 단정 아님)", (s.overbill_unproven_amount||0).toLocaleString()],
    ["누락 (Joint × NDT)", s.underbill_pairs],
    ["분석된 (도면×Joint)", s.joints_analyzed],
    ["도면 요구사항 확보", s.joints_with_required_known],
  ].map(([l,v])=>`<div class="kpi"><div class="label">${esc(l)}</div><div class="value">${esc(v)}</div></div>`).join('');

  // 과다 청구
  buildTable('ob-table',
    ["도면 본체","Joint No.","청구 NDT (과다)","성적서번호","금액","검사일","도면 요구 NDT","billing_item_id"],
    ou.overbill.map(o=>[
      `<span class="code">${esc(o.drawing_body)}</span>`,
      esc(o.joint_no),
      `<span class="badge ke">${esc(o.billed_method)}</span>`,
      esc(o.report_no||''),
      o.amount!=null ? Number(o.amount).toLocaleString() : '−',
      esc(o.inspection_date||''),
      `<span class="small">${esc(o.required_ndt.join(', '))}</span>`,
      String(o.billing_item_id),
    ]), {classes:{0:'code'}});

  bindFilter('ob-table', {
    q:document.getElementById('ob-search'),
    n:document.getElementById('ob-ndt-filter'),
  }, (r, ins)=>{
    if (ins.q.value && !r.textContent.toLowerCase().includes(ins.q.value.toLowerCase())) return false;
    if (ins.n.value && !r.cells[2].textContent.toUpperCase().includes(ins.n.value)) return false;
    return true;
  }, document.getElementById('ob-count'));

  // 근거 미제출·미확정 — 2026-09-05 근거 게이트. 확정과 **다른 표, 다른 금액**. 환불 단정 아님.
  const unp = ou.overbill_unproven || [];
  const unpEl = document.getElementById('ob-unproven');
  if (unpEl) buildTable('ob-unproven',
    ["도면 본체","Joint No.","청구 NDT (도면 밖)","근거 상태","성적서번호","금액","검사일","도면 요구 NDT","billing_item_id"],
    unp.map(o=>[
      `<span class="code">${esc(o.drawing_body)}</span>`,
      esc(o.joint_no),
      `<span class="badge">${esc(o.billed_method)}</span>`,
      esc(o.basis_reason||''),
      esc(o.report_no||''),
      o.amount!=null ? Number(o.amount).toLocaleString() : '−',
      esc(o.inspection_date||''),
      `<span class="small">${esc((o.required_ndt||[]).join(', '))}</span>`,
      String(o.billing_item_id),
    ])
  );

  // 누락
  buildTable('ub-table',
    ["도면 본체","Joint No.","누락 NDT","청구된 NDT","도면 요구 NDT"],
    ou.underbill_joints.map(u=>[
      `<span class="code">${esc(u.drawing_body)}</span>`,
      esc(u.joint_no),
      `<span class="badge warn">${esc(u.missing_method)}</span>`,
      esc(u.billed_methods.join(', ')),
      esc(u.required_ndt.join(', ')),
    ]), {classes:{0:'code'}});

  bindFilter('ub-table', {q:document.getElementById('ub-search')}, (r, ins)=>{
    return !ins.q.value || r.textContent.toLowerCase().includes(ins.q.value.toLowerCase());
  }, document.getElementById('ub-count'));

  // per_joint
  buildTable('pj-table',
    ["도면 본체","Joint No.","도면 요구","청구","과다","누락","상태"],
    ou.per_joint.map(p=>{
      let badge;
      if (p.status === '과다+누락') badge = '<span class="badge ke">과다+누락</span>';
      else if (p.status === '과다') badge = '<span class="badge ke">과다</span>';
      else if (p.status === '누락') badge = '<span class="badge warn">누락</span>';
      else if (p.status.indexOf('미확정') >= 0) badge = '<span class="badge info">미확정</span>';
      else badge = '<span class="badge ok">OK</span>';
      return [
        `<span class="code">${esc(p.drawing_body)}</span>`, esc(p.joint_no),
        esc(p.required_ndt.join(', ')||'−'), esc(p.billed_ndt.join(', ')||'−'),
        esc(p.overbill_methods.join(', ')||''), esc(p.underbill_methods.join(', ')||''),
        badge,
      ];
    }), {classes:{0:'code'}});

  bindFilter('pj-table', {
    q:document.getElementById('pj-search'),
    s:document.getElementById('pj-status-filter'),
  }, (r, ins)=>{
    if (ins.q.value && !r.textContent.toLowerCase().includes(ins.q.value.toLowerCase())) return false;
    if (ins.s.value && r.cells[6].textContent.indexOf(ins.s.value) < 0) return false;
    return true;
  }, document.getElementById('pj-count'));
}

// ─── Safety Matrix ───
function renderSafety(){
  const headers = ["도면","KKS code","재질","Dout×t","Safety Class","Classification","Seismic","QA","Design P (MPag)","Design T (°C)","Op P","Op T","Hydraulic Test P","Min Wall T","Insulation","page","근거 인용 (원문)"];
  const rows = DATA.kks_safety.map(l=>[
    esc(l.drawing_no), esc(l.kks_code), esc(l.material), esc(l.dout_x_thickness_mm),
    l.safety_class_np_001_15 ? `<span class="badge ok">${esc(l.safety_class_np_001_15)}</span>` : '−',
    esc(l.classification_np_001), esc(l.seismic_category_np_031_01), esc(l.qa_category),
    esc(l.design_pressure_mpag), esc(l.design_temperature_c),
    esc(l.operating_pressure_mpag), esc(l.operating_temperature_c),
    esc(l.hydraulic_test_pressure_mpag), esc(l.min_wall_temperature_c),
    esc(l.thermal_insulation), esc(l.page_in_drawing),
    `<span class="cite">${esc(l.citation_quote||'')}</span>`,
  ]);
  buildTable('safety-table', headers, rows, {classes:{1:'kks',16:'cite'}});

  // class filter options
  const classes = Array.from(new Set(DATA.kks_safety.map(l=>l.safety_class_np_001_15).filter(Boolean))).sort();
  const csel = document.getElementById('safety-class-filter');
  classes.forEach(c=>{ const o=document.createElement('option'); o.value=c; o.textContent='Class '+c; csel.appendChild(o); });
  const dwgs = Array.from(new Set(DATA.kks_safety.map(l=>l.drawing_no).filter(Boolean))).sort();
  const dsel = document.getElementById('safety-drawing-filter');
  dwgs.forEach(d=>{ const o=document.createElement('option'); o.value=d; o.textContent=d; dsel.appendChild(o); });

  bindFilter('safety-table', {
    q:document.getElementById('safety-search'),
    c:document.getElementById('safety-class-filter'),
    d:document.getElementById('safety-drawing-filter'),
  }, (r, ins)=>{
    const text = r.textContent.toLowerCase();
    if (ins.q.value && !text.includes(ins.q.value.toLowerCase())) return false;
    if (ins.c.value && !r.cells[4].textContent.includes(ins.c.value)) return false;
    if (ins.d.value && !r.cells[0].textContent.includes(ins.d.value)) return false;
    return true;
  }, document.getElementById('safety-count'));
}

// ─── Inspection Matrix ───
function renderInspection(){
  const headers = ["도면","KKS code","Dout×t","Joint Cat","Safety","VT %","PT/MT %","RT %","UT %","Aux VT %","Aux PT/MT %","page","근거 인용 (원문)"];
  const rows = DATA.inspection_matrix.map(ins=>[
    esc(ins.drawing_no), esc(ins.kks_code), esc(ins.dout_x_thickness_mm), esc(ins.joint_category),
    esc(ins.safety_class_np_001_15),
    fmtPct(ins.vt_pct), fmtPct(ins.pt_or_mt_pct), fmtPct(ins.rt_pct), fmtPct(ins.ut_pct),
    fmtPct(ins.auxiliary_vt_pct), fmtPct(ins.auxiliary_pt_or_mt_pct),
    esc(ins.page_in_drawing),
    `<span class="cite">${esc(ins.citation_quote||'')}</span>`,
  ]);
  buildTable('ins-table', headers, rows, {classes:{1:'kks',12:'cite'}});
  bindFilter('ins-table', {
    q:document.getElementById('ins-search'),
    n:document.getElementById('ins-ndt-filter'),
  }, (r, ins)=>{
    const text = r.textContent.toLowerCase();
    if (ins.q.value && !text.includes(ins.q.value.toLowerCase())) return false;
    if (ins.n.value){
      const colMap = {vt_pct:5, pt_or_mt_pct:6, rt_pct:7, ut_pct:8};
      const v = r.cells[colMap[ins.n.value]].textContent.trim();
      if (!v || v === '−') return false;
    }
    return true;
  }, document.getElementById('ins-count'));
}
function fmtPct(v){
  if (v === null || v === undefined || v === '') return '<span class="dash">−</span>';
  return `<span class="q">${esc(v)}</span>`;
}

// ─── Code Gap ───
function renderCodeGap(){
  const cg = DATA.code_gap;
  const k = document.getElementById('codegap-kpi');
  k.innerHTML = [
    ["⚠ 수령 필요", cg.cited_only.length],
    ["✓ 매칭됨", cg.matched.length],
    ["ℹ 잉여", cg.ingested_only.length],
  ].map(([l,v])=>`<div class="kpi"><div class="label">${esc(l)}</div><div class="value">${v}</div></div>`).join('');

  const n = document.getElementById('codegap-note');
  n.textContent = "도면이 인용한 코드를 NP/PNAE/GOST/TU/ASME/ISO/RCC-M/EN/KS 패밀리별로 분류하고, " +
                  "적재된 표준 문서(standard_documents) 와 정규화 매칭한 결과. 가장 시급한 것은 '수령 필요' 목록.";

  buildTable('codegap-cited-only',
    ["Family","Code (raw)","인용 횟수","인용 도면 수","인용 도면 예시"],
    cg.cited_only.map(c=>[
      `<span class="badge warn">${esc(c.family)}</span>`,
      `<span class="code">${esc(c.raw)}</span>`,
      String(c.count),
      String(c.citing_drawings.length),
      `<span class="small">${esc(c.citing_drawings.slice(0,3).join(', '))}${c.citing_drawings.length>3?' ... +'+(c.citing_drawings.length-3):''}</span>`,
    ]), {classes:{1:'code'}});

  buildTable('codegap-matched',
    ["Family","Code","인용","적재 경로"],
    cg.matched.map(c=>[
      `<span class="badge ok">${esc(c.family)}</span>`,
      `<span class="code">${esc(c.raw)}</span>`,
      String(c.count),
      `<span class="small">${esc(c.ingested_path)}</span>`,
    ]), {classes:{1:'code'}});

  buildTable('codegap-ingested-only',
    ["Family","Code","적재 경로"],
    cg.ingested_only.map(c=>[
      `<span class="badge info">${esc(c.family)}</span>`,
      `<span class="code">${esc(c.raw)}</span>`,
      `<span class="small">${esc(c.ingested_path)}</span>`,
    ]), {classes:{1:'code'}});
}

// ─── Drawing → Codes ───
function renderDrawingCodes(){
  const headers = ["도면번호","인용 표준","수"];
  const rows = Object.entries(DATA.drawing_codes).sort().map(([d, codes])=>[
    `<span class="code">${esc(d)}</span>`,
    codes.map(c=>`<span class="code">${esc(c)}</span>`).join(', '),
    String(codes.length),
  ]);
  buildTable('dc-table', headers, rows);
  bindFilter('dc-table', {q:document.getElementById('dc-search')}, (r, ins)=>{
    return !ins.q.value || r.textContent.toLowerCase().includes(ins.q.value.toLowerCase());
  }, null);
}

// ─── Billing Coverage ───
function renderBillingCoverage(){
  const prefixes = DATA.billing_indexed_prefixes;
  const headers = ["청구 도면번호","참조 건수","DB 상태","비고"];
  const rows = DATA.billing_drawings.map(b=>{
    const dwg = b.drawing_no;
    const body = dwg.split('.DC.')[0].split('.SD.')[0].split('.BG.')[0].split('.KE.')[0];
    const inDb = prefixes.some(p=>p.startsWith(body));
    const isKe = dwg.indexOf('.KE.') >= 0;
    let badge, note;
    if (isKe){ badge = '<span class="badge ke">KE 오기</span>'; note = '시공사 정정 요청'; }
    else if (inDb){ badge = '<span class="badge ok">적재됨</span>'; note = ''; }
    else { badge = '<span class="badge warn">미적재</span>'; note = '도면 수령 필요'; }
    return [`<span class="code">${esc(dwg)}</span>`, String(b.count), badge, esc(note)];
  });
  buildTable('bd-table', headers, rows);
  bindFilter('bd-table', {
    q:document.getElementById('bd-search'),
    s:document.getElementById('bd-status-filter'),
  }, (r, ins)=>{
    if (ins.q.value && !r.textContent.toLowerCase().includes(ins.q.value.toLowerCase())) return false;
    if (ins.s.value){
      const c = r.cells[2].textContent;
      if (ins.s.value === 'missing' && !c.includes('미적재')) return false;
      if (ins.s.value === 'indexed' && !c.includes('적재됨')) return false;
      if (ins.s.value === 'ke' && !c.includes('KE')) return false;
    }
    return true;
  }, document.getElementById('bd-count'));
}

// ─── Standards Catalog ───
function renderStandards(){
  const headers = ["종류","문서번호","Rev","적재일","경로"];
  const rows = DATA.standards.map(s=>[
    `<span class="badge info">${esc(s.doc_type)}</span>`,
    `<span class="code">${esc(s.document_no)}</span>`,
    esc(s.revision || '−'),
    esc(s.ingested_at || ''),
    `<span class="small">${esc(s.file_path)}</span>`,
  ]);
  buildTable('std-table', headers, rows);
}

// ─── Nav highlight on scroll ───
function setupNavHighlight(){
  const links = document.querySelectorAll('#topnav a');
  const sections = SECTIONS.map(([id])=>document.getElementById(id));
  window.addEventListener('scroll', ()=>{
    let active = 0;
    sections.forEach((s,i)=>{ if (s && s.getBoundingClientRect().top < 100) active = i; });
    links.forEach((l,i)=> l.classList.toggle('active', i===active));
  });
}

renderNav();
renderKPIs();
renderOverbill();
renderSafety();
renderInspection();
renderCodeGap();
renderDrawingCodes();
renderBillingCoverage();
renderStandards();
setupNavHighlight();
</script>
</body>
</html>
"""


def _write_md(snap: dict, path: Path) -> None:
    lines: list[str] = []
    lines.append(f"# 검사기준 가이드 — {snap['discipline']}\n")
    lines.append(f"_생성일: {snap['generated_at']}_\n")

    lines.append("## 1. 개요\n")
    lines.append(f"- 도면 세트: **{len(snap['drawing_sets'])}** 개")
    lines.append(f"- KKS code (Safety Class 표 행): **{len(snap['kks_safety'])}** 개")
    lines.append(f"- 검사 매트릭스 행: **{len(snap['inspection_matrix'])}** 개")
    lines.append(f"- 적재된 표준 문서: **{len(snap['standards'])}** 개")
    lines.append(f"- 청구에 등장하는 고유 도면: **{len(snap['billing_drawings'])}** 개\n")

    lines.append("## 2. Safety Class 분포\n")
    sc_dist = collections.Counter(
        (line.get("safety_class_np_001_15") or "(unspecified)") for line in snap["kks_safety"]
    )
    if sc_dist:
        lines.append("| Safety Class | KKS code 수 |")
        lines.append("|---|---|")
        for sc, n in sorted(sc_dist.items(), key=lambda x: str(x[0])):
            lines.append(f"| {sc} | {n} |")
        lines.append("")

    lines.append("## 3. 도면별 적용 표준 인용\n")
    for dwg, codes in sorted(snap["drawing_codes"].items()):
        lines.append(f"- **{dwg}** → {', '.join(sorted(codes))}")
    lines.append("")

    lines.append("## 4. KKS code 별 검사 요구사항 표본 (상위 20)\n")
    if snap["inspection_matrix"]:
        lines.append("| 도면 | KKS code | Dout×t | VT% | PT/MT% | RT% | UT% |")
        lines.append("|---|---|---|---|---|---|---|")
        for ins in snap["inspection_matrix"][:20]:
            lines.append(
                f"| {ins.get('drawing_no')} | {ins.get('kks_code')} | "
                f"{ins.get('dout_x_thickness_mm') or '-'} | "
                f"{ins.get('vt_pct') if ins.get('vt_pct') is not None else '-'} | "
                f"{ins.get('pt_or_mt_pct') if ins.get('pt_or_mt_pct') is not None else '-'} | "
                f"{ins.get('rt_pct') if ins.get('rt_pct') is not None else '-'} | "
                f"{ins.get('ut_pct') if ins.get('ut_pct') is not None else '-'} |"
            )
        if len(snap["inspection_matrix"]) > 20:
            lines.append(f"\n_(... +{len(snap['inspection_matrix']) - 20} 행)_")
    else:
        lines.append("⚠ 적재된 도면에서 추출된 검사 매트릭스가 없습니다. "
                     "도면을 더 적재하거나 DC parser 의 표 추출이 동작했는지 확인하세요.")
    lines.append("")

    lines.append("## 5. 적재된 표준 문서 카탈로그\n")
    if snap["standards"]:
        lines.append("| 종류 | 문서번호 | Rev. | 경로 |")
        lines.append("|---|---|---|---|")
        for std in sorted(snap["standards"], key=lambda s: (s.doc_type, s.file_path)):
            lines.append(f"| {std.doc_type} | {std.document_no} | {std.revision or '-'} | `{std.file_path}` |")
    else:
        lines.append("⚠ 적재된 표준 문서가 없습니다. "
                     "`python -m app.main ingest-standards <폴더> --type {scwep|code|contract}` 로 적재하세요.")
    lines.append("")

    lines.append("## 6. 청구 도면 커버리지\n")
    indexed = {ds.drawing_no for ds in snap["drawing_sets"]}
    missing: list[tuple] = []
    ke_misentry: list[tuple] = []
    for dwg, n in snap["billing_drawings"].most_common():
        body = dwg.split(".DC.")[0].split(".SD.")[0].split(".BG.")[0].split(".KE.")[0]
        if ".KE." in dwg:
            ke_misentry.append((dwg, n))
        elif not any(idx.startswith(body) for idx in indexed):
            missing.append((dwg, n))
    lines.append(f"- 적재된 도면 본체: **{len(indexed)}**")
    lines.append(f"- 청구가 참조하나 도면 DB 에 없음: **{len(missing)}** 개")
    lines.append(f"- KE(WEP/SCWEP) 오기로 분류된 청구 도면: **{len(ke_misentry)}** 개")
    if missing:
        lines.append("\n### 도면 DB 에 미적재 — 수령 필요\n")
        for dwg, n in missing[:20]:
            lines.append(f"- ({n} 건 참조) `{dwg}`")
        if len(missing) > 20:
            lines.append(f"- _(... +{len(missing) - 20} 개)_")
    lines.append("")

    lines.append("## 7. 검토자 사용 안내\n")
    lines.append(
        "- **검토 회의용**: 시공사·NIS 와 협의 시 `Safety Class 매트릭스`·`검사 요구 매트릭스` 시트를 "
        "근거 자료로 직접 인용 가능. 모든 행에 도면번호 + KKS code + 원문 인용 동봉.\n"
        "- **불일치 발견 시**: 청구 엑셀의 KKS code 또는 Joint No. 를 이 가이드의 매트릭스와 대조해 "
        "요구사항 누락·과다 청구를 정량 확인.\n"
        "- **KE 오기**: 청구 도면번호에 `.KE.` 가 들어있으면 시공사 정정 요청 대상 (자동 finding 발생).\n"
        "- **갱신**: 신규 도면 적재 시 `python -m app.main criteria-guide --discipline CP-P1` 재실행."
    )

    path.write_text("\n".join(lines), encoding="utf-8")
