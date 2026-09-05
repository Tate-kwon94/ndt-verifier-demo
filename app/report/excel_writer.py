"""검토용 엑셀 출력 — 원본 청구 엑셀에 검토 컬럼 추가.

색상 코딩: 위험도에 따라 행 배경색 적용.
"""
from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select

from app.config import matching_rules
from app.database.models import BillingItem, BillingRound, Finding, InspectionReport, Match, get_session

logger = logging.getLogger(__name__)

# 2026-09-05 (SCWEP 근거 게이트):
#  · "근거_상태" 신설 — 판정 **바로 옆**. 상태는 판정과 같이 읽혀야 한다. 검토자_메모는 계속 맨 끝.
#  · "누락_검사" → "타행_요구NDT" 로 개명. 값(still_required)이 **행 단위**라서, VT+RT 를 정상 청구한
#    Joint 의 VT 행에 "누락: RT" 가 찍혔다. 그대로 두면 검토자가 이미 청구된 검사를 시공사에 추궁한다.
#  · 감사에서 잡힌 "항상 빈칸" 3개(도면_요구NDT·타행_요구NDT·과다_검사)는 아래 도우미가 실제 키를 읽도록 고쳤다.
REVIEW_COLUMNS = [
    "매칭_성적서번호",
    "매칭_방법",
    "매칭_점수",
    "도면_요구NDT",
    "타행_요구NDT",
    "과다_검사",
    "적합성_판정",
    "근거_상태",
    "위험도",
    "재확인_필요",
    "재확인_사유",
    "LLM_설명",
    "근거_문서",
    "근거_조항/페이지",
    "권장_조치",
    "검토자_메모",
]


def write(billing_round_id: int, output_path: Optional[Path] = None) -> Path:
    with get_session() as s:
        br = s.get(BillingRound, billing_round_id)
        if br is None:
            raise ValueError(f"BillingRound {billing_round_id} 미존재")
        src = Path(br.billing_xlsx_path)
        if not src.exists():
            raise FileNotFoundError(f"원본 청구 엑셀 누락: {src}")

        items = s.scalars(
            select(BillingItem).where(BillingItem.billing_round_id == billing_round_id)
        ).all()
        item_by_id = {i.id: i for i in items}

        matches = {m.billing_item_id: m for m in s.scalars(select(Match))}
        findings = {f.billing_item_id: f for f in s.scalars(select(Finding))}
        reports = {r.id: r for r in s.scalars(select(InspectionReport))}

    wb = openpyxl.load_workbook(src)

    # 분야(br.discipline) 의 columns 만 candidate_tokens 로 → threshold 가 적정해짐
    from app.extractors.excel_parser import (
        _detect_discipline_by_name,
        _find_header_row_and_cells,
        _normalize_token,
    )
    from app.config import templates_config

    cfg = templates_config()["disciplines"]
    discipline_cfg = cfg.get(br.discipline) or {}
    candidate_tokens: set[str] = set()
    for col in discipline_cfg.get("columns", []) or []:
        for c in col.get("candidates", []):
            candidate_tokens.add(_normalize_token(c))

    chosen_ws = None
    chosen_header_row = 1
    best_rows = -1
    for sname in wb.sheetnames:
        if _detect_discipline_by_name(sname) != br.discipline:
            continue
        ws_try = wb[sname]
        hr, _hc = _find_header_row_and_cells(
            ws_try, discipline_cfg.get("header_row_search_max", 10), candidate_tokens
        )
        if hr is None:
            continue
        # 데이터 행 수로 main 시트 결정
        data_rows = max(0, ws_try.max_row - hr)
        if data_rows > best_rows:
            chosen_ws = ws_try
            chosen_header_row = hr
            best_rows = data_rows

    ws = chosen_ws or wb.active

    # 기존 끝 컬럼 다음에 리뷰 컬럼 추가 (헤더는 원본 헤더 행에 맞춤)
    start_col = ws.max_column + 1
    for offset, header in enumerate(REVIEW_COLUMNS):
        cell = ws.cell(row=chosen_header_row, column=start_col + offset, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    highlight_threshold = int(matching_rules().get("review_priority", {}).get("highlight_threshold", 60))

    # 원본 엑셀 행 → BillingItem 매핑 (parser 가 raw_json["_excel_row"] 에 저장)
    items_by_excel_row = {}
    for i in items:
        er = (i.raw_json or {}).get("_excel_row")
        if er:
            items_by_excel_row[int(er)] = i

    for excel_row in range(chosen_header_row + 1, ws.max_row + 1):
        item = items_by_excel_row.get(excel_row)
        if item is None:
            continue
        m = matches.get(item.id)
        f = findings.get(item.id)
        rep = reports.get(m.inspection_report_id) if (m and m.inspection_report_id) else None

        values = _row_values(item, m, f, rep)
        for offset, val in enumerate(values):
            cell = ws.cell(row=excel_row, column=start_col + offset, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 색상 코딩: 재확인 필요(파랑계열) vs 고위험(빨강계열) 분리
        # 재확인 필요가 위험도보다 우선 (검토자의 시선이 가장 먼저 닿아야 함)
        needs_review = _aggregate_needs_review(m, f, rep)
        if needs_review:
            color = "BBDEFB"   # 연파랑 — 재확인
            for col in range(1, start_col + len(REVIEW_COLUMNS)):
                ws.cell(row=excel_row, column=col).fill = PatternFill("solid", fgColor=color)
        elif f and f.risk_score >= highlight_threshold:
            color = _risk_color(f.risk_score)
            for col in range(1, start_col + len(REVIEW_COLUMNS)):
                ws.cell(row=excel_row, column=col).fill = PatternFill("solid", fgColor=color)

    output_path = output_path or (src.parent / f"{src.stem}_검토결과.xlsx")
    wb.save(output_path)
    return output_path


# ─────────────────────────── helpers ───────────────────────────


def _row_values(item, m, f, rep) -> list:
    drawing_required = _required_ndt_summary(f)
    missing = _missing_methods(f)
    extra = _extra_methods(f)
    citations_summary = _citations_summary(f)
    review_reasons = _collect_review_reasons(m, f, rep)

    return [
        rep.report_no if rep else None,
        m.match_method if m else None,
        round(m.match_score, 1) if (m and m.match_score is not None) else None,
        drawing_required,
        ", ".join(missing) if missing else None,
        ", ".join(extra) if extra else None,
        f.verdict if f else None,
        _basis_state_label(f),
        f.risk_score if f else None,
        "필요" if review_reasons else None,
        "\n".join(f"- {r}" for r in review_reasons) if review_reasons else None,
        f.summary if f else None,
        citations_summary[0] if citations_summary else None,
        citations_summary[1] if citations_summary else None,
        f.recommended_action if f else None,
        None,  # 검토자_메모 — 사람이 채움
    ]


def _aggregate_needs_review(m, f, rep) -> bool:
    return (
        (f is not None and f.needs_review)
        or (m is not None and m.needs_review)
        or (rep is not None and rep.needs_review)
    )


def _collect_review_reasons(m, f, rep) -> list[str]:
    reasons: list[str] = []
    if f and f.needs_review:
        for r in ((f.review_reasons_json or {}).get("reasons") or []):
            reasons.append(f"[적합성] {r}")
    if m and m.needs_review:
        for r in ((m.review_reasons_json or {}).get("reasons") or []):
            reasons.append(f"[매칭] {r}")
    if rep and rep.needs_review:
        for r in ((rep.review_reasons_json or {}).get("reasons") or []):
            reasons.append(f"[성적서] {r}")
    return reasons


# 과다청구 계열 rule 4개 — 근거 게이트 이후 "도면에 없는 검사를 청구했다" 는 이 넷 중 하나로 나온다.
_OVERBILL_FAMILY = ("billed_ndt_not_in_requirements", "billed_ndt_basis_not_submitted",
                    "billed_ndt_basis_unclear", "billed_ndt_covered_by_scwep")

_BASIS_LABEL = {
    "covered": "SCWEP 근거 있음 — 조건 확인 필요",
    "unclear": "근거 불명확 — 조항 확인",
    "not_submitted": "근거 미제출 — 제출 요청",
    "no_basis_found": "근거 밖 — 과다청구",
}


def _findings_of(f: Optional[Finding]) -> list[dict]:
    return list(((f.citations_json or {}).get("findings") if f else None) or [])


def _overbill_finding(f: Optional[Finding]) -> Optional[dict]:
    for fe in _findings_of(f):
        if fe.get("rule") in _OVERBILL_FAMILY:
            return fe
    return None


def _basis_state_label(f: Optional[Finding]) -> Optional[str]:
    fe = _overbill_finding(f)
    if not fe:
        return None
    return _BASIS_LABEL.get((fe.get("details") or {}).get("basis_state"))


def _required_ndt_summary(f: Optional[Finding]) -> Optional[str]:
    """도면이 이 Joint 에 요구한 NDT. 4단계부터 citations_json['drawing_requirement'] 에 실려 온다.

    예전 코드는 explanation_json['findings_explained'][*]['details']['missing_methods'] 를 읽었는데
    프롬프트는 그 키를 만들지 않는다 — 이 칸은 지금까지 항상 빈칸이었다 (감사 2026-09-04).
    """
    if f is None:
        return None
    dr = (f.citations_json or {}).get("drawing_requirement") or {}
    items = ((dr.get("required_ndt_json") or {}).get("items") if isinstance(dr, dict) else None) or []
    methods = sorted({str(i.get("method")).strip().upper() for i in items
                      if isinstance(i, dict) and i.get("method")})
    if methods:
        return ", ".join(methods)
    # 도면 요구가 실려 있지 않으면 과다청구 finding 이 갖고 있는 요구 목록으로
    fe = _overbill_finding(f)
    if fe:
        req = (fe.get("details") or {}).get("required_ndt_by_drawing") or []
        return ", ".join(req) if req else None
    return None


def _missing_methods(f: Optional[Finding]) -> list[str]:
    """이 행 말고 다른 행에서 청구됐어야 할 도면 요구 NDT (행 단위 — 그래서 컬럼명이 '타행_요구NDT').

    예전 키 'missing_methods' 는 compliance 가 만든 적이 없다. 실제 키는 'still_required'.
    """
    for fe in _findings_of(f):
        if fe.get("rule") == "required_ndt_missing":
            return list((fe.get("details") or {}).get("still_required") or [])
    return []


def _extra_methods(f: Optional[Finding]) -> list[str]:
    """도면에 없는데 청구된 NDT. 예전 키 'requested' 는 없었다 — 실제 키는 'requested_ndt'.
    근거 게이트 이후엔 확정(not_in_requirements)뿐 아니라 미제출·불명확·근거있음 행도 같은 값을 갖는다 —
    '도면 밖 청구' 라는 사실은 넷 다 같고, 근거_상태 칸이 그 뜻을 가른다."""
    fe = _overbill_finding(f)
    if fe:
        v = (fe.get("details") or {}).get("requested_ndt")
        return [v] if v else []
    return []


def _citations_summary(f: Optional[Finding]):
    """근거 문서 / 조항·쪽. LLM 이 되읊은 인용이 먼저, 없으면 결정론 authority_refs.

    예전엔 explanation_json 만 읽어 SCWEP 결정론 인용이 이 칸에 영원히 못 들어갔다.
    단, **과다청구 확정(no_basis_found) 행에는 SCWEP 인용을 찍지 않는다** — 고발 옆에 SCWEP 이 붙으면
    그 SCWEP 이 청구를 인정한 것처럼 읽힌다.
    """
    if f is None:
        return None
    docs: list[str] = []
    locs: list[str] = []

    def _take(c: dict) -> None:
        doc = c.get("doc")
        if doc:
            docs.append(str(doc))
        loc = []
        if c.get("section"):
            loc.append(str(c["section"]))
        if c.get("page") is not None:
            loc.append(f"p.{c['page']}")
        if loc:
            locs.append("/".join(loc))

    expl = f.explanation_json or {}
    for fe in expl.get("findings_explained", []) or []:
        for c in fe.get("evidence_citations", []) or []:
            _take(c)
    if not docs:
        ob = _overbill_finding(f)
        accused = bool(ob) and (ob.get("details") or {}).get("basis_state") == "no_basis_found"
        for r in (f.citations_json or {}).get("authority_refs") or []:
            if accused and r.get("authority_level") == 2:
                continue
            _take(r)
    if not docs and not locs:
        return None
    return (" / ".join(sorted(set(docs))) or None, " / ".join(sorted(set(locs))) or None)


def _risk_color(score: int) -> str:
    # 빨강·주황·노랑 그라데이션
    if score >= 80:
        return "FFCCCC"
    if score >= 60:
        return "FFE0B2"
    return "FFF9C4"
